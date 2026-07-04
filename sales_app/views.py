import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from django.shortcuts import render
from .models import Sales
from django.db.models import Sum, Count, Avg, FloatField, ExpressionWrapper, F, Q, Min,OuterRef,Max,Case,IntegerField, Sum, When, DecimalField
from django.db.models.functions import ExtractMonth, ExtractDay, TruncDay, ExtractWeek
from django.db.models.functions import TruncMonth, TruncWeek, ExtractHour, ExtractWeekDay
from .models import Sales, UserProfile
from django.http import HttpResponse
from datetime import datetime, date, timedelta
from django.utils import timezone
from django.http import JsonResponse
import os
from django.conf import settings
import calendar
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from django.db import connection
from django.contrib import messages
from itertools import combinations
from dateutil.relativedelta import relativedelta
from collections import defaultdict
import re

from django.db.models.functions import Left
from django.db.models import Count, Case, When, Value, IntegerField
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden

from django.contrib.auth import login, logout ,authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import render, redirect
from django.db.models import Prefetch

from sales_app.decorators import cache_dashboard_view

from django.contrib.admin.views.decorators import staff_member_required
import math
from django.utils.safestring import mark_safe

# At the top of your views.py, outside any view
def calculate_filter_options(current_year, selected_locations, selected_category, selected_product, selected_campaign, user_profile):
    """Shared logic for calculating available filter options"""
    allowed_locations = user_profile.get_allowed_locations()
    
    base_query = Sales.objects.filter(cd__year=current_year).exclude(
        un__in=["მთავარი საწყობი 2", "სატესტო"]
    )
    
    # LOCATIONS
    q_locations = base_query
    if selected_category != 'all':
        q_locations = q_locations.filter(prodg=selected_category)
    if selected_product != 'all':
        q_locations = q_locations.filter(prod=selected_product)
    if selected_campaign != 'all':
        q_locations = q_locations.filter(actions=selected_campaign)
    available_locations = list(q_locations.values_list('un', flat=True).distinct().order_by('un'))
    
    # CATEGORIES
    q_categories = base_query
    if selected_locations:
        q_categories = q_categories.filter(un__in=selected_locations)
    if selected_product != 'all':
        q_categories = q_categories.filter(prod=selected_product)
    if selected_campaign != 'all':
        q_categories = q_categories.filter(actions=selected_campaign)
    available_categories = list(q_categories.values_list('prodg', flat=True).distinct().order_by('prodg'))
    
    # PRODUCTS
    q_products = base_query
    if selected_locations:
        q_products = q_products.filter(un__in=selected_locations)
    if selected_category != 'all':
        q_products = q_products.filter(prodg=selected_category)
    if selected_campaign != 'all':
        q_products = q_products.filter(actions=selected_campaign)
    available_products = list(q_products.values_list('prod', flat=True).distinct().order_by('prod'))
    
    # CAMPAIGNS
    q_campaigns = base_query
    if selected_locations:
        q_campaigns = q_campaigns.filter(un__in=selected_locations)
    if selected_category != 'all':
        q_campaigns = q_campaigns.filter(prodg=selected_category)
    if selected_product != 'all':
        q_campaigns = q_campaigns.filter(prod=selected_product)
    available_campaigns = list(q_campaigns.values_list('actions', flat=True).distinct().order_by('actions'))
    
    return {
        'locations': available_locations if user_profile.is_admin else allowed_locations,
        'categories': available_categories,
        'products': available_products,
        'campaigns': available_campaigns
    }

def user_login(request):
    # If already logged in, go to dashboard
    if request.user.is_authenticated:
        return redirect('sales_dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Check if user has profile
            try:
                profile = user.profile
                login(request, user)
                messages.success(request, f'Welcome back, {user.username}!')
                
                # Try to redirect to 'next' parameter, otherwise dashboard
                next_url = request.GET.get('next', 'another')
                return redirect(next_url)
                
            except Exception as e:
                print(f"Profile error: {e}")  # Debug
                messages.error(request, "Your account is not configured. Contact administrator.")
                return redirect('login')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'login.html')

def user_logout(request):
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('login')
    
# @cache_dashboard_view(timeout=900)
@login_required
def dashboard(request):
    """Optimized dashboard view with reduced queries and better performance"""
    
    # ==================== SETUP & VALIDATION ====================
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        return HttpResponseForbidden("Access denied. Contact administrator.")
    
    allowed_locations = user_profile.get_allowed_locations()
    # Year selection
    comparison_mode = request.GET.get('comparison', '2026-2025')
    if comparison_mode == '2026-2025':
        current_year, previous_year = 2026, 2025
    elif comparison_mode == '2026-2024':
        current_year, previous_year = 2026, 2024
    else:
        current_year, previous_year = 2025, 2024
    
    # Date parsing
    start_date_str = request.GET.get('start_date', f'{current_year}-07-01')
    end_date_str = request.GET.get('end_date', f'{current_year}-07-31')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except:
        start_date = date(current_year, 1, 1)
    
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except:
        end_date = date(current_year, 12, 31)
    
    # Location security check - UPDATED FOR ADMIN PERFORMANCE
    selected_locations = request.GET.getlist('un_filter')
    
    if not user_profile.is_admin:
        # Non-admin logic (unchanged)
        if not selected_locations or 'all' in selected_locations:
            selected_locations = allowed_locations
        else:
            unauthorized = set(selected_locations) - set(allowed_locations)
            if unauthorized:
                messages.warning(request, f"Access denied to: {', '.join(unauthorized)}")
                selected_locations = [loc for loc in selected_locations if loc in allowed_locations]
            if not selected_locations:
                selected_locations = allowed_locations
    else:
        # ADMIN: Default to top location to avoid slow "all" query
        if not selected_locations:
            # First visit - auto-select top location by revenue
            top_location = Sales.objects.filter(
                cd__year=current_year
            ).values('un').annotate(
                total=Sum('tanxa')
            ).order_by('-total').values_list('un', flat=True).first()
            
            if top_location:
                selected_locations = [top_location]
                messages.info(request, f"Showing data for {top_location}. Use filters to view other locations or all data.")
            else:
                selected_locations = []
        elif 'all' in selected_locations:
            # Explicitly selected "all" - allow but warn about performance
            selected_locations = []
            messages.warning(request, "Loading all locations - this may take 1-2 minutes. Consider selecting specific locations for faster results.")
    
    if not selected_locations and not user_profile.is_admin:
        return HttpResponseForbidden("You don't have access to any locations. Contact administrator.")
    
    # Display selected location
    if user_profile.is_admin and (not selected_locations or 'all' in request.GET.getlist('un_filter')):
        selected_un = 'all'
    else:
        selected_un = selected_locations[0] if len(selected_locations) == 1 else 'multiple'
    
    # Other filters
    selected_category = request.GET.get('category', 'all')
    selected_product = request.GET.get('prod_filter', 'all')
    selected_campaign = request.GET.get('campaign_filter', 'all')
    
    # Adjust dates to current year
    start_date = start_date.replace(year=current_year)
    end_date = end_date.replace(year=current_year)
    
    # Get max date for current year (OPTIMIZED - single query)
    max_date_query = Sales.objects.filter(cd__year=current_year)
    if selected_locations:
        max_date_query = max_date_query.filter(un__in=selected_locations)
    if selected_category != 'all':
        max_date_query = max_date_query.filter(prodg=selected_category)
    if selected_product != 'all':
        max_date_query = max_date_query.filter(prod=selected_product)
    if selected_campaign != 'all':
        max_date_query = max_date_query.filter(actions=selected_campaign)
    
    max_date = max_date_query.aggregate(max_date=Max('cd'))['max_date']
    
    if max_date and end_date > max_date.date():
        end_date = max_date.date()
    
    # Previous year dates
    previous_start = start_date.replace(year=previous_year)
    previous_end = end_date.replace(year=previous_year)
    
    # Timezone-aware datetimes
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    previous_start_datetime = timezone.make_aware(datetime.combine(previous_start, datetime.min.time()))
    previous_end_datetime = timezone.make_aware(datetime.combine(previous_end, datetime.max.time()))
    
    date_filter_current = {
        'cd__year': current_year,
        'cd__gte': start_datetime,
        'cd__lte': end_datetime
    }
    
    date_filter_previous = {
        'cd__year': previous_year,
        'cd__gte': previous_start_datetime,
        'cd__lte': previous_end_datetime
    }
    
    # ==================== HELPER FUNCTIONS ====================
    
    def calc_change(current, previous):
        if previous and previous > 0:
            return ((current - previous) / previous) * 100
        return 0

    def apply_filters(q):
        """Apply all filters consistently"""
        if selected_locations:
            q = q.filter(un__in=selected_locations)
        if selected_category != 'all':
            q = q.filter(prodg=selected_category)
        if selected_product != 'all':
            q = q.filter(prod=selected_product)
        if selected_campaign != 'all':
            q = q.filter(actions=selected_campaign)
        return q
    
    def get_base_queryset(is_current=True):
        """Get base queryset with filters applied"""
        if is_current:
            q = Sales.objects.filter(**date_filter_current).exclude(
                un__in=["მთავარი საწყობი 2", "სატესტო"]
            )
        else:
            q = Sales.objects.filter(**date_filter_previous).exclude(
                un__in=["მთავარი საწყობი 2", "სატესტო"]
            )
        return apply_filters(q)
    
    def get_comprehensive_stats(is_current=True):
        """
        PRODUCTION OPTIMIZED: Prevents timeouts with proper filtering
        """
        # Use the pre-built date filters from parent scope
        if is_current:
            base_filter = date_filter_current
        else:
            base_filter = date_filter_previous
        
        # Build query using the complete filter
        q = Sales.objects.filter(**base_filter).exclude(
            un__in=["მთავარი საწყობი 2", "სატესტო"]
        )
        
        if selected_locations:
            q = q.filter(un__in=selected_locations)
        if selected_category != 'all':
            q = q.filter(prodg=selected_category)
        if selected_product != 'all':
            q = q.filter(prod=selected_product)
        if selected_campaign != 'all':
            q = q.filter(actions=selected_campaign)
        
        # STEP 1: Get simple aggregates (NO DISTINCT - very fast)
        try:
            totals = q.aggregate(
                    total_revenue=Sum('tanxa'),
                    total_items=Sum('raod'),
                    discount_total=Sum('discount_price'),
                    std_price_total=Sum('std_price'),
                    skincare=Sum(
                                Case(
                                    When(prodg='SKIN CARE', then='tanxa'),
                                    default=0,
                                    output_field=FloatField()
                                )
                            ),
                    total_without_pop=Sum(
                            Case(
                                When(~Q(prodg='POP'), then='tanxa'),  # ← was summing POP, now excludes it
                                default=0,
                                output_field=FloatField()
                            )
                        ),
                    # count rows where tanxa == std_price (sold at full price)
                    full_price_count=Sum(
                        Case(
                            When(tanxa=F('std_price'), then=1),
                            default=0,
                            output_field=IntegerField()
                        )
                    ),

                    # sum of tanxa only for those rows
                    full_price_revenue=Sum(
                        Case(
                            When(tanxa=F('std_price'), then=F('tanxa')),
                            default=0,
                            output_field=DecimalField()
                        )
                    ),
                )
        except Exception as e:
            print(f"❌ Stats query failed: {e}")
            return {
                'daily_data': [],
                'total_revenue': 0,
                'total_tickets': 0,
                'total_items': 0,
                'avg_basket': 0,
                'discount_share': 0,
                'identification_rate': 0,    # ← add
                'skincare_percentage': 0,    # ← add
            }
        
        ticket_stats = q.aggregate(
            total_tickets=Count('zedd', distinct=True),
            anonymous_tickets=Count(
                Case(
                    When(Gvari__startswith='ფიზიკური პირი', then='zedd'),
                    output_field=IntegerField()
                ),
                distinct=True
            )
        )

        total_tickets        = ticket_stats['total_tickets']
        anonymous_tickets    = ticket_stats['anonymous_tickets']
        identification_rate = (
            round((1 - (anonymous_tickets / total_tickets)) * 100, 1)
            if total_tickets > 0 else 0
        )
        # STEP 3: Simplified daily data with ticket counts per day
        try:
            # Get daily revenue and items
            daily_data = list(
                q.annotate(
                    month=ExtractMonth('cd'),
                    day=ExtractDay('cd')
                ).values('month', 'day').annotate(
                    revenue=Sum('tanxa'),
                    items=Sum('raod'),
                    discount_total=Sum('discount_price'),
                    std_price_total=Sum('std_price'),
                    tickets=Count('zedd', distinct=True)  # FIXED: Get actual daily ticket count
                ).order_by('month', 'day')[:366]
            )
        except Exception as e:
            print(f"❌ Daily data failed: {e}")
            daily_data = []
        
        # Calculate final metrics
        total_revenue = float(totals['total_revenue'] or 0)
        total_items = totals['total_items'] or 0
        total_discount = float(totals['discount_total'] or 0)
        total_std_price = float(totals['std_price_total'] or 0)
        skincare_without_pop = float(totals['total_without_pop'] or 0)
        skincare_revenue = float(totals['skincare'] or 0)
        skincare_percentage = (skincare_revenue / skincare_without_pop * 100) if skincare_without_pop > 0 else 0
        
        avg_basket = total_revenue / total_tickets if total_tickets > 0 else 0
        discount_share = (1 - (total_discount / total_std_price)) * 100 if total_std_price > 0 else 0
        
        return {
            'daily_data': daily_data,
            'total_revenue': total_revenue,
            'total_tickets': total_tickets,
            'total_items': total_items,
            'avg_basket': avg_basket,
            'discount_share': discount_share,
            'identification_rate': identification_rate,
            'skincare_percentage': skincare_percentage
        }

    def get_cross_selling_stats(is_current=True):
        """
        OPTIMIZED: Single query with conditional aggregation for cross-selling
        """
        if is_current:
            date_filter = date_filter_current
        else:
            date_filter = date_filter_previous
        
        q = (Sales.objects
            .filter(**date_filter, prodt='selling item')
            .exclude(tanxa=0)
            .exclude(prodg='POP')
            .exclude(idprod__in=['M9157', 'M9121', 'M9850']))
        
        q = apply_filters(q)
        
        # Single query to get ticket-level counts
        tickets_with_counts = q.values('zedd').annotate(
            item_count=Count('idreal1')
        )
        
        # Convert to list once
        ticket_list = list(tickets_with_counts)
        
        total_tickets = len(ticket_list)
        if total_tickets == 0:
            return {
                'cross_sell_tickets': 0,
                'cross_sell_percentage': 0,
                'single_item_tickets': 0,
                'single_item_percentage': 0,
                'total_tickets': 0
            }
        
        cross_sell_tickets = sum(1 for t in ticket_list if t['item_count'] >= 3)
        single_item_tickets = sum(1 for t in ticket_list if t['item_count'] == 1)
        
        return {
            'cross_sell_tickets': cross_sell_tickets,
            'cross_sell_percentage': (cross_sell_tickets / total_tickets * 100),
            'single_item_tickets': single_item_tickets,
            'single_item_percentage': (single_item_tickets / total_tickets * 100),
            'total_tickets': total_tickets
        }
    
    def get_daily_cross_selling_stats(is_current=True):
        """
        OPTIMIZED: Get daily cross-selling percentages
        """
        if is_current:
            date_filter = date_filter_current
        else:
            date_filter = date_filter_previous
        
        q = Sales.objects.filter(**date_filter, prodt='selling item').exclude(tanxa=0).exclude(prodg='POP')
        q = apply_filters(q)
        
        # Get all data in one query
        daily_data = q.annotate(
            month=ExtractMonth('cd'),
            day=ExtractDay('cd')
        ).values('month', 'day', 'zedd').annotate(
            item_count=Count('idreal1')
        )
        
        # Process in Python (faster than multiple DB queries)
        date_stats = {}
        for record in daily_data:
            date_key = f"{record['month']}/{record['day']}"
            if date_key not in date_stats:
                date_stats[date_key] = {'total': 0, 'single_item': 0, 'cross_sell': 0}
            
            date_stats[date_key]['total'] += 1
            if record['item_count'] == 1:
                date_stats[date_key]['single_item'] += 1
            elif record['item_count'] >= 3:
                date_stats[date_key]['cross_sell'] += 1
        
        # Convert to percentages
        result = {}
        for date_key, stats in date_stats.items():
            total = stats['total']
            result[date_key] = {
                'single_item_pct': (stats['single_item'] / total * 100) if total > 0 else 0,
                'cross_sell_pct': (stats['cross_sell'] / total * 100) if total > 0 else 0,
                'total_tickets': total
            }
        
        return result
    
    def get_ticket_distribution(is_current=True):
        """
        OPTIMIZED: Get ticket amount distribution
        """
        if is_current:
            date_filter = date_filter_current
        else:
            date_filter = date_filter_previous
        
        q = Sales.objects.filter(**date_filter, prodt='selling item').exclude(tanxa=0)
        q = apply_filters(q)
        
        # Single query to get ticket totals
        ticket_totals = list(q.values('zedd').annotate(
            ticket_total=Sum('tanxa')
        ).values_list('ticket_total', flat=True))
        
        if not ticket_totals:
            return {
                'distribution': {},
                'distribution_pct': {},
                'total_tickets': 0,
                'avg_ticket': 0,
                'median_ticket': 0,
                'p25': 0,
                'p75': 0
            }
        
        # Define ranges
        ranges = [
            (0, 50, '0-50'), (50, 100, '50-100'), (100, 150, '100-150'),
            (150, 200, '150-200'), (200, 300, '200-300'), (300, 500, '300-500'),
            (500, 1000, '500-1K'), (1000, float('inf'), '1K+')
        ]
        
        distribution = {label: 0 for _, _, label in ranges}
        total_tickets = len(ticket_totals)
        
        # Categorize tickets
        for amount in ticket_totals:
            amount = float(amount)
            for min_val, max_val, label in ranges:
                if min_val <= amount < max_val:
                    distribution[label] += 1
                    break
        
        # Calculate percentages
        distribution_pct = {
            label: (count / total_tickets * 100) if total_tickets > 0 else 0
            for label, count in distribution.items()
        }
        
        # Statistics
        ticket_list = [float(t) for t in ticket_totals]
        sorted_tickets = sorted(ticket_list)
        
        return {
            'distribution': distribution,
            'distribution_pct': distribution_pct,
            'total_tickets': total_tickets,
            'avg_ticket': sum(ticket_list) / len(ticket_list),
            'median_ticket': sorted_tickets[len(sorted_tickets) // 2],
            'p25': sorted_tickets[len(sorted_tickets) // 4],
            'p75': sorted_tickets[3 * len(sorted_tickets) // 4]
        }
    
    def get_product_analysis(is_current=True):
        """
        OPTIMIZED: Get product performance with smart limiting
        FIXED: Added avg_ticket_value calculation
        """
        q = get_base_queryset(is_current).exclude(prodg='POP')
        
        # Only get top 100 products by revenue (not ALL products)
        products = list(
            q.values('prod', 'idprod')
            .annotate(
                total_revenue=Sum('tanxa'),
                quantity=Sum('raod'),
                tickets=Count('zedd', distinct=True),
                last_purchase_date=Max('cd')
            )
            .order_by('-total_revenue')[:100]  # LIMIT to top 100
        )
        
        if not products:
            return {
                'bestsellers': [],
                'least_sellers': [],
                'slow_movers': [],
                'rising_stars': []
            }
        
        # Calculate avg_ticket_value for each product (total_revenue / tickets)
        for product in products:
            product['avg_ticket_value'] = (
                float(product['total_revenue'] or 0) / product['tickets'] 
                if product['tickets'] > 0 else 0
            )
        
        # Calculate performance scores
        max_revenue = max(p['total_revenue'] for p in products)
        max_frequency = max(p['tickets'] for p in products)
        max_monetary = max(p['avg_ticket_value'] for p in products if p['avg_ticket_value'])
        
        for product in products:
            # Recency score
            if product['last_purchase_date']:
                last_purchase = product['last_purchase_date']
                if timezone.is_naive(last_purchase):
                    last_purchase = timezone.make_aware(last_purchase)
                days_since = (end_datetime - last_purchase).days
                product['recency_days'] = days_since
                product['recency_score'] = max(0, 100 - days_since)
            else:
                product['recency_days'] = 999
                product['recency_score'] = 0
            
            product['revenue'] = float(product['total_revenue'] or 0)
            
            # Normalized scores
            revenue_normalized = (product['revenue'] / max_revenue * 100) if max_revenue > 0 else 0
            frequency_normalized = (product['tickets'] / max_frequency * 100) if max_frequency > 0 else 0
            monetary_normalized = (product['avg_ticket_value'] / max_monetary * 100) if max_monetary > 0 else 0
            
            # Composite score
            product['performance_score'] = (
                revenue_normalized * 0.40 +
                frequency_normalized * 0.30 +
                product['recency_score'] * 0.20 +
                monetary_normalized * 0.10
            )
            
            # Tier classification
            if product['performance_score'] >= 80:
                product['tier'] = 'S'
                product['tier_label'] = 'Top Performer'
            elif product['performance_score'] >= 60:
                product['tier'] = 'A'
                product['tier_label'] = 'Strong Seller'
            elif product['performance_score'] >= 40:
                product['tier'] = 'B'
                product['tier_label'] = 'Average'
            elif product['performance_score'] >= 20:
                product['tier'] = 'C'
                product['tier_label'] = 'Weak Seller'
            else:
                product['tier'] = 'D'
                product['tier_label'] = 'Poor Performer'
        
        # Sort and categorize
        products_sorted = sorted(products, key=lambda x: x['performance_score'], reverse=True)
        
        return {
            'bestsellers': products_sorted[:15],
            'least_sellers': sorted(products, key=lambda x: x['performance_score'])[:15],
            'slow_movers': sorted([p for p in products if p['recency_days'] > 30], 
                                 key=lambda x: x['recency_days'], reverse=True)[:10],
            'rising_stars': sorted([p for p in products if p['recency_score'] > 70], 
                                  key=lambda x: (x['recency_score'], x['tickets']), reverse=True)[:10]
        }
    
    def get_association_rules():
        q = (Sales.objects
            .filter(**date_filter_current, prodt='selling item')
            .exclude(tanxa=0)
            .exclude(prodg='POP')
        )
        
        if selected_locations:
            q = q.filter(un__in=selected_locations)
        if selected_category != 'all':
            q = q.filter(prodg=selected_category)
        
        # Single query - get ticket + product pairs
        pairs = q.values('zedd', 'prod')
        
        # Build transactions
        transactions = defaultdict(set)
        for row in pairs:
            if row['prod']:
                transactions[row['zedd']].add(row['prod'])
        
        transaction_list = list(transactions.values())
        total_transactions = len(transaction_list)
        
        if total_transactions == 0:
            return []
        
        # Count frequencies
        product_freq = defaultdict(int)
        pair_freq = defaultdict(int)
        
        for basket in transaction_list:
            for product in basket:
                product_freq[product] += 1
            for pair in combinations(sorted(basket), 2):
                pair_freq[pair] += 1
        
        # Generate rules
        MIN_SUPPORT = 0.01
        MIN_CONFIDENCE = 0.10
        MIN_LIFT = 1.2
        
        rules = []
        for (item_a, item_b), pair_count in pair_freq.items():
            support = pair_count / total_transactions
            if support < MIN_SUPPORT:
                continue
            
            for antecedent, consequent in [(item_a, item_b), (item_b, item_a)]:
                confidence = pair_count / product_freq[antecedent]
                lift = confidence / (product_freq[consequent] / total_transactions)
                
                if confidence >= MIN_CONFIDENCE and lift >= MIN_LIFT:
                    rules.append({
                        'antecedent': antecedent,
                        'consequent': consequent,
                        'support': round(support * 100, 2),
                        'confidence': round(confidence * 100, 1),
                        'lift': round(lift, 2),
                        'pair_count': pair_count,
                    })
        
        rules.sort(key=lambda x: x['lift'], reverse=True)
        return rules[:50]
        
    def get_campaign_analytics():
        """
        Enhanced campaign analysis:
        - YoY comparison per campaign
        - Revenue share %
        - Effectiveness vs baseline (no-campaign avg basket)
        - Monthly trend data for top campaigns
        """
        # from django.db.models import Q

        q_curr = get_base_queryset(is_current=True)
        q_prev = get_base_queryset(is_current=False)

        # ---- Top 15 campaigns from current year ----
        top_campaigns_curr = list(
            q_curr
            .exclude(Q(actions__isnull=True) | Q(actions=''))
            .values('actions')
            .annotate(
                revenue=Sum('tanxa'),
                tickets=Count('zedd', distinct=True),
                quantity=Sum('raod'),
                avg_basket=ExpressionWrapper(
                    Sum('tanxa') / Count('zedd', distinct=True),
                    output_field=FloatField()
                ),
                discount_total=Sum('discount_price'),
                std_price_total=Sum('std_price'),
            )
            .order_by('-revenue')[:15]
        )

        campaign_names = [c['actions'] for c in top_campaigns_curr]

        # ---- Previous year data for same campaigns ----
        top_campaigns_prev = list(
            q_prev
            .filter(actions__in=campaign_names)
            .values('actions')
            .annotate(
                revenue=Sum('tanxa'),
                tickets=Count('zedd', distinct=True),
                avg_basket=ExpressionWrapper(
                    Sum('tanxa') / Count('zedd', distinct=True),
                    output_field=FloatField()
                )
            )
        )
        prev_dict = {c['actions']: c for c in top_campaigns_prev}

        # ---- Monthly trend for top 5 campaigns ----
        top_5_names = campaign_names[:5]
        monthly_campaign_data = list(
            q_curr
            .filter(actions__in=top_5_names)
            .annotate(month=ExtractMonth('cd'))
            .values('month', 'actions')
            .annotate(revenue=Sum('tanxa'))
            .order_by('month')
        )

        # Build monthly trend dict: {campaign: [rev_jan, rev_feb, ...]}
        campaign_monthly = {name: [0] * 12 for name in top_5_names}
        for row in monthly_campaign_data:
            if row['actions'] in campaign_monthly:
                campaign_monthly[row['actions']][row['month'] - 1] = float(row['revenue'] or 0)

        # ---- No-campaign baseline ----
        no_camp_q = get_base_queryset(is_current=True).filter(
            Q(actions__isnull=True) | Q(actions='')
        )
        no_camp_stats = no_camp_q.aggregate(
            revenue=Sum('tanxa'),
            tickets=Count('zedd', distinct=True)
        )
        no_camp_revenue = float(no_camp_stats['revenue'] or 0)
        no_camp_tickets = no_camp_stats['tickets'] or 0
        no_camp_avg_basket = no_camp_revenue / no_camp_tickets if no_camp_tickets > 0 else 0

        # ---- Build comparison list ----
        total_rev = stats_current['total_revenue']
        campaign_total_revenue = sum(float(c['revenue'] or 0) for c in top_campaigns_curr)

        campaign_comparison = []
        for camp in top_campaigns_curr:
            prev = prev_dict.get(camp['actions'], {})
            rev_curr = float(camp['revenue'] or 0)
            rev_prev = float(prev.get('revenue') or 0)
            tickets_curr = camp['tickets'] or 0
            tickets_prev = prev.get('tickets') or 0
            avg_basket_curr = float(camp['avg_basket'] or 0)
            avg_basket_prev = float(prev.get('avg_basket') or 0)
            discount_total = float(camp['discount_total'] or 0)
            std_price_total = float(camp['std_price_total'] or 0)
            discount_share_camp = (1 - (discount_total / std_price_total)) * 100 if std_price_total > 0 else 0

            rev_change = calc_change(rev_curr, rev_prev)
            basket_change_camp = calc_change(avg_basket_curr, avg_basket_prev)
            tickets_change_camp = calc_change(tickets_curr, tickets_prev)
            rev_share = (rev_curr / total_rev * 100) if total_rev > 0 else 0

            # Effectiveness: how much higher/lower is this campaign's avg basket vs overall
            effectiveness = (avg_basket_curr / avg_basket_current) if avg_basket_current > 0 else 1
            # vs no-campaign
            vs_no_campaign = (avg_basket_curr / no_camp_avg_basket) if no_camp_avg_basket > 0 else 1

            campaign_comparison.append({
                'name': camp['actions'],
                'revenue_current': rev_curr,
                'revenue_previous': rev_prev,
                'tickets_current': tickets_curr,
                'tickets_previous': tickets_prev,
                'avg_basket_current': avg_basket_curr,
                'avg_basket_previous': avg_basket_prev,
                'rev_change': rev_change,
                'basket_change': basket_change_camp,
                'tickets_change': tickets_change_camp,
                'rev_share': rev_share,
                'effectiveness': effectiveness,
                'vs_no_campaign': vs_no_campaign,
                'quantity': camp['quantity'] or 0,
                'discount_share': discount_share_camp,
            })

        return {
            'campaign_comparison': campaign_comparison,
            'no_camp_revenue': no_camp_revenue,
            'no_camp_tickets': no_camp_tickets,
            'no_camp_avg_basket': no_camp_avg_basket,
            'active_campaigns': len(top_campaigns_curr),
            'campaign_total_revenue': campaign_total_revenue,
            'campaign_revenue_share': (campaign_total_revenue / total_rev * 100) if total_rev > 0 else 0,
            'campaign_monthly': campaign_monthly,
            'top_5_campaign_names': top_5_names,
        }
    
    def get_location_summary():
        """
        Per-location breakdown table:
        Columns: Location | Turnover (curr / % vs prev) | Tickets | Identification Rate
                Avg Basket | Median Basket | Cross-sell Rate | Single Item Rate
        """
        from django.db.models import FloatField

        # ---- Current year per-location aggregates ----
        q_curr = (
            Sales.objects
            .filter(**date_filter_current)
            .exclude(un__in=["მთავარი საწყობი 2", "სატესტო"])
        )
        if selected_locations:
            q_curr = q_curr.filter(un__in=selected_locations)
        if selected_category != 'all':
            q_curr = q_curr.filter(prodg=selected_category)
        if selected_product != 'all':
            q_curr = q_curr.filter(prod=selected_product)
        if selected_campaign != 'all':
            q_curr = q_curr.filter(actions=selected_campaign)

        # ---- Previous year per-location aggregates ----
        q_prev = (
            Sales.objects
            .filter(**date_filter_previous)
            .exclude(un__in=["მთავარი საწყობი 2", "სატესტო"])
        )
        if selected_locations:
            q_prev = q_prev.filter(un__in=selected_locations)
        if selected_category != 'all':
            q_prev = q_prev.filter(prodg=selected_category)
        if selected_product != 'all':
            q_prev = q_prev.filter(prod=selected_product)
        if selected_campaign != 'all':
            q_prev = q_prev.filter(actions=selected_campaign)

        # ── Revenue, tickets, items per location ──────────────────────────────
        curr_agg = list(
            q_curr.values('un').annotate(
                revenue=Sum('tanxa'),
                tickets=Count('zedd', distinct=True),
                items=Sum('raod'),
                discount_total=Sum('discount_price'),
                std_price_total=Sum('std_price'),
                anon_tickets=Count(
                    Case(
                        When(Gvari__startswith='ფიზიკური პირი', then='zedd'),
                        output_field=IntegerField()
                    ),
                    distinct=True
                ),
                skincare=Sum(
                    Case(
                        When(prodg='SKIN CARE', then='tanxa'),
                        default=0,
                        output_field=FloatField()
                    )
                ) / Sum(
                    Case(
                        When(~Q(prodg='POP'), then='tanxa'),
                        default=0,
                        output_field=FloatField()
                    )
                )
            )
        )

        prev_agg = list(
        q_prev.values('un').annotate(
            revenue=Sum('tanxa'),
            tickets=Count('zedd', distinct=True),
            items=Sum('raod'),
            anon_tickets=Count(
                Case(
                    When(Gvari__startswith='ფიზიკური პირი', then='zedd'),
                    output_field=IntegerField()
                ),
                distinct=True
            ),
            discount_total=Sum('discount_price'),
            std_price_total=Sum('std_price'),
            skincare=Sum(
                    Case(
                        When(prodg='SKIN CARE', then='tanxa'),
                        default=0,
                        output_field=FloatField()
                    )
                ) / Sum(
                    Case(
                        When(~Q(prodg='POP'), then='tanxa'),
                        default=0,
                        output_field=FloatField()
                    )
                )
            )
        )

        # ── Per-location cross-sell stats (current) ───────────────────────────
        q = q_curr.exclude(tanxa=0).exclude(prodg='POP')
        q1 = q_prev.exclude(tanxa=0).exclude(prodg='POP')

        # Group by location + ticket to get item count per ticket per location
        tickets_with_counts = list(
            q.values('un', 'zedd').annotate(item_count=Count('idreal1'))
        )
        tickets_with_counts1 = list(
            q1.values('un', 'zedd').annotate(item_count=Count('idreal1'))
        )

        # Build per-location dicts: { un: { cross_sell, single_item, total } }
        def build_cross_sell_dict(ticket_rows):
            d = {}
            for t in ticket_rows:
                un = t['un']
                if un not in d:
                    d[un] = {'cross_sell': 0, 'single_item': 0, 'total': 0}
                d[un]['total'] += 1
                if t['item_count'] >= 3:
                    d[un]['cross_sell'] += 1
                if t['item_count'] == 1:
                    d[un]['single_item'] += 1
            return d

        cross_sell_curr = build_cross_sell_dict(tickets_with_counts)
        cross_sell_prev = build_cross_sell_dict(tickets_with_counts1)

        # ── Median basket per location (current) ──────────────────────────────
        ticket_totals_by_loc = {}
        ticket_rows = list(
            q_curr.filter(prodt='selling item')
            .exclude(tanxa=0)
            .values('un', 'zedd')
            .annotate(ticket_total=Sum('tanxa'))
        )
        for row in ticket_rows:
            loc = row['un']
            if loc not in ticket_totals_by_loc:
                ticket_totals_by_loc[loc] = []
            ticket_totals_by_loc[loc].append(float(row['ticket_total'] or 0))

        def median_of(lst):
            if not lst:
                return 0
            s = sorted(lst)
            n = len(s)
            return s[n // 2] if n % 2 == 1 else (s[n // 2 - 1] + s[n // 2]) / 2

        # ── Build lookup dicts ─────────────────────────────────────────────────
        prev_dict = {r['un']: r for r in prev_agg}

        rows = []
        for loc in curr_agg:
            un = loc['un']
            rev_curr   = float(loc['revenue'] or 0)
            skincare_curr = float(loc['skincare'] or 0) * 100  # ← add * 100
            tix_curr   = loc['tickets'] or 0
            items_curr = loc['items'] or 0
            anon       = loc['anon_tickets'] or 0
            disc_tot   = float(loc['discount_total'] or 0)
            std_tot    = float(loc['std_price_total'] or 0)

            prev       = prev_dict.get(un, {})
            skincare_prev = float(prev.get('skincare') or 0) * 100  # ← add * 100
            rev_prev   = float(prev.get('revenue') or 0)
            tix_prev   = prev.get('tickets') or 0
            items_prev = prev.get('items') or 0

            avg_basket_curr = rev_curr / tix_curr if tix_curr > 0 else 0
            avg_basket_prev = rev_prev / tix_prev if tix_prev > 0 else 0

            ticket_vals   = ticket_totals_by_loc.get(un, [])
            median_basket = median_of(ticket_vals)

            anon_prev      = prev.get('anon_tickets') or 0
            disc_tot_prev  = float(prev.get('discount_total') or 0)
            std_tot_prev   = float(prev.get('std_price_total') or 0)

            id_rate_prev       = round((1 - (anon_prev / tix_prev)) * 100, 2) if tix_prev > 0 else 0
            discount_share_prev = round((1 - (disc_tot_prev / std_tot_prev)) * 100, 2) if std_tot_prev > 0 else 0
            id_rate        = round((1 - (anon / tix_curr)) * 100, 2) if tix_curr > 0 else 0
            discount_share = round((1 - (disc_tot / std_tot)) * 100, 2) if std_tot > 0 else 0

            # Cross-sell rates per location
            cs_curr  = cross_sell_curr.get(un, {})
            cs_prev  = cross_sell_prev.get(un, {})
            cs_total_curr = cs_curr.get('total', 0)
            cs_total_prev = cs_prev.get('total', 0)

            cross_sell_rate_curr  = round(cs_curr.get('cross_sell', 0)  / cs_total_curr * 100, 2) if cs_total_curr > 0 else 0
            single_item_rate_curr = round(cs_curr.get('single_item', 0) / cs_total_curr * 100, 2) if cs_total_curr > 0 else 0
            cross_sell_rate_prev  = round(cs_prev.get('cross_sell', 0)  / cs_total_prev * 100, 2) if cs_total_prev > 0 else 0
            single_item_rate_prev = round(cs_prev.get('single_item', 0) / cs_total_prev * 100, 2) if cs_total_prev > 0 else 0

            rows.append({
                'un':                   un,
                'rev_curr':             rev_curr,
                'rev_prev':             rev_prev,
                'rev_chg':              calc_change(rev_curr, rev_prev),
                'tix_curr':             tix_curr,
                'tix_prev':             tix_prev,
                'tix_chg':              calc_change(tix_curr, tix_prev),
                'items_curr':           items_curr,
                'items_prev':           items_prev,
                'items_chg':            calc_change(items_curr, items_prev),
                'avg_basket_curr':      avg_basket_curr,
                'avg_basket_prev':      avg_basket_prev,
                'avg_basket_chg':       calc_change(avg_basket_curr, avg_basket_prev),
                'median_basket':        median_basket,
                'id_rate':              id_rate,
                'discount_share':       discount_share,
                'cross_sell_rate':      cross_sell_rate_curr,
                'cross_sell_rate_prev': cross_sell_rate_prev,
                'cross_sell_chg':       calc_change(cross_sell_rate_curr, cross_sell_rate_prev),
                'single_item_rate':     single_item_rate_curr,
                'single_item_rate_prev':single_item_rate_prev,
                'single_item_chg':      calc_change(single_item_rate_curr, single_item_rate_prev),
                'id_rate_prev':         id_rate_prev,
                'id_rate_chg':          calc_change(id_rate, id_rate_prev),
                'discount_share_prev':  discount_share_prev,
                'discount_share_chg':   calc_change(discount_share, discount_share_prev),
                'skincare_curr':        skincare_curr,
                'skincare_prev':        skincare_prev,
                'skincare_change':        calc_change(skincare_curr, skincare_prev),
            })

        # Sort by current revenue descending
        rows.sort(key=lambda x: x['rev_curr'], reverse=True)
        return rows

    location_summary = get_location_summary()
    association_rules = get_association_rules()
    # Get comprehensive stats for both years (2 queries instead of many)
    stats_current = get_comprehensive_stats(is_current=True)
    stats_previous = get_comprehensive_stats(is_current=False)
    # campaign_analytics = get_campaign_analytics()
    
    # Cross-selling stats (2 queries)
    cross_sell_current = get_cross_selling_stats(is_current=True)
    cross_sell_previous = get_cross_selling_stats(is_current=False)
    
    # Daily cross-selling (2 queries)
    cross_sell_daily_current = get_daily_cross_selling_stats(is_current=True)
    cross_sell_daily_previous = get_daily_cross_selling_stats(is_current=False)
    
    # Ticket distribution (2 queries)
    dist_current = get_ticket_distribution(is_current=True)
    dist_previous = get_ticket_distribution(is_current=False)
    
    # Product analysis (2 queries)
    product_analysis_current = get_product_analysis(is_current=True)
    
    # ==================== PREPARE CHART DATA ====================
    
    # Extract daily data
    data_current = stats_current['daily_data']
    data_previous = stats_previous['daily_data']
    
    # Create date maps
    date_map_revenue_current = {f"{i['month']}/{i['day']}": float(i['revenue'] or 0) for i in data_current}
    date_map_revenue_previous = {f"{i['month']}/{i['day']}": float(i['revenue'] or 0) for i in data_previous}
    
    date_map_tickets_current = {f"{i['month']}/{i['day']}": int(i['tickets'] or 0) for i in data_current}
    date_map_tickets_previous = {f"{i['month']}/{i['day']}": int(i['tickets'] or 0) for i in data_previous}
    
    date_map_items_current = {f"{i['month']}/{i['day']}": int(i['items'] or 0) for i in data_current}
    date_map_items_previous = {f"{i['month']}/{i['day']}": int(i['items'] or 0) for i in data_previous}
    
    # Generate labels
    labels = [f"{i['month']}/{i['day']}" for i in data_current]
    
    # Map values
    values_current = [date_map_revenue_current.get(label, 0) for label in labels]
    values_previous = [date_map_revenue_previous.get(label, 0) for label in labels]
    
    tickets_values_current = [date_map_tickets_current.get(label, 0) for label in labels]
    tickets_values_previous = [date_map_tickets_previous.get(label, 0) for label in labels]
    
    items_values_current = [date_map_items_current.get(label, 0) for label in labels]
    items_values_previous = [date_map_items_previous.get(label, 0) for label in labels]
    
    # Cross-selling arrays
    single_item_pct_current = [cross_sell_daily_current.get(label, {}).get('single_item_pct', 0) for label in labels]
    single_item_pct_previous = [cross_sell_daily_previous.get(label, {}).get('single_item_pct', 0) for label in labels]
    
    cross_sell_pct_current = [cross_sell_daily_current.get(label, {}).get('cross_sell_pct', 0) for label in labels]
    cross_sell_pct_previous = [cross_sell_daily_previous.get(label, {}).get('cross_sell_pct', 0) for label in labels]
    
    # Calculate average basket per day - FIXED LOGIC
    basket_values_current = []
    basket_values_previous = []
    
    for label in labels:
        revenue_curr = date_map_revenue_current.get(label, 0)
        tickets_curr = date_map_tickets_current.get(label, 0)
        basket_values_current.append(revenue_curr / tickets_curr if tickets_curr > 0 else 0)
        
        revenue_prev = date_map_revenue_previous.get(label, 0)
        tickets_prev = date_map_tickets_previous.get(label, 0)
        basket_values_previous.append(revenue_prev / tickets_prev if tickets_prev > 0 else 0)
    
    # ==================== CALCULATE METRICS & CHANGES ====================
    
    total_current = stats_current['total_revenue']
    total_previous = stats_previous['total_revenue']
    total_tickets_current = stats_current['total_tickets']
    total_tickets_previous = stats_previous['total_tickets']
    total_items_current = stats_current['total_items']
    total_items_previous = stats_previous['total_items']
    avg_basket_current = stats_current['avg_basket']
    avg_basket_previous = stats_previous['avg_basket']
    discount_share_current = stats_current['discount_share']
    discount_share_previous = stats_previous.get('discount_share', 0)
    identification_rate_current = stats_current['identification_rate']
    identification_rate_previous = stats_previous.get('identification_rate', 0)
    skincare_percentage_current = stats_current.get('skincare_percentage', 0)   # ← parentheses
    skincare_percentage_previous = stats_previous.get('skincare_percentage', 0)  # ← parentheses
    campaign_analytics = get_campaign_analytics()
    
    def calc_change(current, previous):
        if previous and previous > 0:
            return ((current - previous) / previous) * 100
        return 0
    
    percentage_change = calc_change(total_current, total_previous)
    tickets_change = calc_change(total_tickets_current, total_tickets_previous)
    items_change = calc_change(total_items_current, total_items_previous)
    basket_change = calc_change(avg_basket_current, avg_basket_previous)
    discount_share_change = calc_change(discount_share_current, discount_share_previous)
    identification_rate_change = calc_change(identification_rate_current, identification_rate_previous)
    skincare_change = calc_change(skincare_percentage_current, skincare_percentage_previous)
    
    cross_sell_change = calc_change(
        cross_sell_current['cross_sell_percentage'],
        cross_sell_previous['cross_sell_percentage']
    )
    
    single_item_change = calc_change(
        cross_sell_current['single_item_percentage'],
        cross_sell_previous['single_item_percentage']
    )
    
    dist_avg_change = calc_change(dist_current['avg_ticket'], dist_previous['avg_ticket'])
    dist_median_change = calc_change(dist_current['median_ticket'], dist_previous['median_ticket'])
    
    # Distribution data for charts
    distribution_labels = ['0-50', '50-100', '100-150', '150-200', '200-300', '300-500', '500-1K', '1K+']
    distribution_counts_current = [dist_current['distribution'].get(label, 0) for label in distribution_labels]
    distribution_counts_previous = [dist_previous['distribution'].get(label, 0) for label in distribution_labels]
    distribution_pct_current = [dist_current['distribution_pct'].get(label, 0) for label in distribution_labels]
    distribution_pct_previous = [dist_previous['distribution_pct'].get(label, 0) for label in distribution_labels]
    
    # Conversion rate
    conversion_rate_current = (total_tickets_current / total_items_current * 100) if total_items_current > 0 else 0
    conversion_rate_previous = (total_tickets_previous / total_items_previous * 100) if total_items_previous > 0 else 0
    conversion_change = calc_change(conversion_rate_current, conversion_rate_previous)
    
    # Active locations
    active_locations_current = get_base_queryset(is_current=True).values('un').distinct().count()
    active_locations_previous = get_base_queryset(is_current=False).values('un').distinct().count()
    locations_change = calc_change(active_locations_current, active_locations_previous)
    # Monthly tickets (2 queries)
    monthly_tickets_current = list(
        get_base_queryset(is_current=True)
        .annotate(month=ExtractMonth('cd'))
        .values('month')
        .annotate(tickets=Count('zedd', distinct=True))
        .order_by('month')
    )
    
    monthly_tickets_previous = list(
        get_base_queryset(is_current=False)
        .annotate(month=ExtractMonth('cd'))
        .values('month')
        .annotate(tickets=Count('zedd', distinct=True))
        .order_by('month')
    )
    
    month_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    tickets_data_current = [0] * 12
    tickets_data_previous = [0] * 12
    
    for item in monthly_tickets_current:
        tickets_data_current[item['month'] - 1] = item['tickets']
    
    for item in monthly_tickets_previous:
        tickets_data_previous[item['month'] - 1] = item['tickets']
    
    # Monthly basket
    monthly_basket_current = list(
        get_base_queryset(is_current=True)
        .annotate(month=ExtractMonth('cd'))
        .values('month')
        .annotate(
            total_revenue=Sum('tanxa'),
            total_tickets=Count('zedd', distinct=True)
        )
        .order_by('month')
    )
    
    basket_data_current = [0] * 12
    for item in monthly_basket_current:
        if item['total_tickets'] and item['total_tickets'] > 0:
            basket_data_current[item['month'] - 1] = float(item['total_revenue'] or 0) / item['total_tickets']
    
    # Category data (1 query)
    category_data = list(
        get_base_queryset(is_current=True)
        .values('prodg')
        .annotate(total=Sum('tanxa'))
        .order_by('-total')[:8]
    )
    
    category_labels = [item['prodg'] or 'Unknown' for item in category_data]
    category_values = [float(item['total'] or 0) for item in category_data]
    
    # Category comparison (2 queries)
    category_query_current_comp = get_base_queryset(is_current=True)
    category_query_previous_comp = get_base_queryset(is_current=False)
    
    category_data_current_comp = list(
        category_query_current_comp
        .values('prodg')
        .annotate(total=Sum('tanxa'))
        .order_by('-total')[:10]
    )
    
    top_categories = [item['prodg'] for item in category_data_current_comp]
    category_data_previous_comp = list(
        category_query_previous_comp
        .filter(prodg__in=top_categories)
        .values('prodg')
        .annotate(total=Sum('tanxa'))
    )
    
    cat_previous_dict = {item['prodg']: float(item['total'] or 0) for item in category_data_previous_comp}
    
    category_comparison = []
    for item in category_data_current_comp:
        cat_name = item['prodg'] or 'Unknown'
        revenue_current = float(item['total'] or 0)
        revenue_previous = cat_previous_dict.get(item['prodg'], 0)
        
        change = revenue_current - revenue_previous
        pct_change = calc_change(revenue_current, revenue_previous)
        
        category_comparison.append({
            'name': cat_name,
            'revenue_previous': revenue_previous,
            'revenue_current': revenue_current,
            'change': change,
            'pct_change': pct_change
        })
    
    # Top 10 Campaigns by value (1 query)
    top_10_campaigns = list(
        get_base_queryset(is_current=True)
        .values('actions')
        .annotate(
            total=Sum('tanxa'),
            quantity=Sum('raod'),
            zedd_unique=Count('zedd', distinct=True),
            avg_basket=ExpressionWrapper(
                Sum('tanxa') / Count('zedd', distinct=True),
                output_field=FloatField()
            )
        )
        .order_by('-total')[:10]
    )

    

    # ==================== FORMATTING HELPERS ====================
    
    def format_currency(value):
        if value >= 1000000:
            return f"₾{value/1000000:.1f}M"
        elif value >= 1000:
            return f"₾{value/1000:.1f}K"
        return f"₾{value:.2f}"
    
    def format_number(value):
        if value >= 1000000:
            return f"{value/1000000:.1f}M"
        elif value >= 1000:
            return f"{value/1000:.1f}K"
        return f"{int(value)}"

    # ==================== GET FILTER OPTIONS ====================
    filter_options = calculate_filter_options(
        current_year, selected_locations, selected_category, 
        selected_product, selected_campaign, user_profile
    )

    all_locations = filter_options['locations']
    all_categories = filter_options['categories']
    all_products = filter_options['products']
    all_campaigns = filter_options['campaigns']
    date_range_text = f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d')}, {current_year}"
    
    # ==================== BUILD CONTEXT ====================
    
    context = {
        'comparison_mode': comparison_mode,
        'current_year': current_year,
        'previous_year': previous_year,
        'max_date': max_date,
        'date_range_text': date_range_text,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        
        # Chart data (JSON)
        'labels': json.dumps(labels),
        'data_previous': json.dumps(values_previous),
        'data_current': json.dumps(values_current),
        'tickets_values_current': json.dumps(tickets_values_current),
        'tickets_values_previous': json.dumps(tickets_values_previous),
        'items_values_current': json.dumps(items_values_current),
        'items_values_previous': json.dumps(items_values_previous),
        'basket_values_current': json.dumps(basket_values_current),
        'basket_values_previous': json.dumps(basket_values_previous),
        'single_item_pct_current': json.dumps(single_item_pct_current),
        'single_item_pct_previous': json.dumps(single_item_pct_previous),
        'cross_sell_pct_current': json.dumps(cross_sell_pct_current),
        'cross_sell_pct_previous': json.dumps(cross_sell_pct_previous),
        'discount_share_previous': json.dumps(discount_share_previous),
        'discount_share_current': json.dumps(discount_share_current),
        # 'client_identification': client_identification_rate,
        # 'client_identification_change': client_identification_change,
        'identification_rate_current':json.dumps(identification_rate_current),
        'identification_rate_previous':json.dumps(identification_rate_previous),
        'identification_rate_change': identification_rate_change,
        'skincare_percentage_current' : json.dumps(skincare_percentage_current),
        'skincare_percentage_previous' : json.dumps(skincare_percentage_previous),
        'skincare_percentage_change' : skincare_change,
        
        'month_labels': json.dumps(month_labels),
        'tickets_data_previous': json.dumps(tickets_data_previous),
        'tickets_data_current': json.dumps(tickets_data_current),
        'basket_data_current': json.dumps(basket_data_current),
        
        'category_labels': json.dumps(category_labels),
        'category_values': json.dumps(category_values),
        'category_comparison': category_comparison,
        
        # Formatted metrics
        'total_current': format_currency(total_current),
        'total_previous': format_currency(total_previous),
        'total_tickets': format_number(total_tickets_current),
        'total_items': format_number(total_items_current),
        'avg_basket': f"₾{avg_basket_current:.2f}",
        'conversion_rate': conversion_rate_current,
        'active_customers': format_number(active_locations_current),
        
        'cross_sell_percentage_current': cross_sell_current['cross_sell_percentage'],
        'cross_sell_percentage_previous': cross_sell_previous['cross_sell_percentage'],
        'cross_sell_change': cross_sell_change,
        'single_item_percentage_current': cross_sell_current['single_item_percentage'],
        'single_item_percentage_previous': cross_sell_previous['single_item_percentage'],
        'single_item_change': single_item_change,
        
        # Changes
        'percentage_change': percentage_change,
        'association_rules': association_rules,
        'tickets_change': tickets_change,
        'basket_change': basket_change,
        'items_change': items_change,
        'conversion_change': conversion_change,
        'customers_change': locations_change,
        'discount_share_precentage_change': discount_share_change,

        'location_summary': location_summary,
        'location_summary_json': json.dumps([
            {**r, 'un': r['un']} for r in location_summary
        ], default=float),

        # campaign
        'campaign_top_10': campaign_analytics['campaign_comparison'][:10],
        'campaign_comparison': campaign_analytics['campaign_comparison'],
        'campaign_analytics': {
            'no_camp_revenue': campaign_analytics['no_camp_revenue'],
            'no_camp_tickets': campaign_analytics['no_camp_tickets'],
            'no_camp_avg_basket': campaign_analytics['no_camp_avg_basket'],
            'active_campaigns': campaign_analytics['active_campaigns'],
            'campaign_total_revenue': campaign_analytics['campaign_total_revenue'],
            'campaign_revenue_share': campaign_analytics['campaign_revenue_share'],
        },
        'campaign_monthly': json.dumps(campaign_analytics['campaign_monthly']),
        'top_5_campaign_names': json.dumps(campaign_analytics['top_5_campaign_names']),
        'no_camp_avg_basket': campaign_analytics['no_camp_avg_basket'],
        'active_campaigns': campaign_analytics['active_campaigns'],
        'campaign_revenue_share': campaign_analytics['campaign_revenue_share'],
        'selected_campaign': selected_campaign,

        # Pre-serialized campaign chart arrays for JS
        'campaign_names_json': json.dumps([c['name'] for c in campaign_analytics['campaign_comparison']]),
        'camp_rev_curr_json': json.dumps([round(c['revenue_current'], 2) for c in campaign_analytics['campaign_comparison']]),
        'camp_rev_prev_json': json.dumps([round(c['revenue_previous'], 2) for c in campaign_analytics['campaign_comparison']]),
        'camp_avg_baskets_json': json.dumps([round(c['avg_basket_current'], 2) for c in campaign_analytics['campaign_comparison']]),
        'camp_tickets_json': json.dumps([c['tickets_current'] for c in campaign_analytics['campaign_comparison']]),
        
        # Other data
        'prod_dt': product_analysis_current['bestsellers'][:10],  # Top 10 only
        
        # Filters
        'all_locations': all_locations,
        'all_categories': all_categories,
        'all_campaigns': all_campaigns,
        'selected_un': selected_un,
        'selected_locations': selected_locations,
        'selected_category': selected_category,
        'selected_product': selected_product,
        'products': all_products,
        # 'high_zedd': top_10_zedd,
        
        # Distribution
        'distribution_labels': json.dumps(distribution_labels),
        'distribution_counts_current': json.dumps(distribution_counts_current),
        'distribution_counts_previous': json.dumps(distribution_counts_previous),
        'distribution_pct_current': json.dumps(distribution_pct_current),
        'distribution_pct_previous': json.dumps(distribution_pct_previous),
        
        'dist_avg_current': dist_current['avg_ticket'],
        'dist_avg_previous': dist_previous['avg_ticket'],
        'dist_avg_change': dist_avg_change,
        'dist_median_current': dist_current['median_ticket'],
        'dist_median_previous': dist_previous['median_ticket'],
        'dist_median_change': dist_median_change,
        'dist_p25_current': dist_current['p25'],
        'dist_p75_current': dist_current['p75'],
        'dist_total_tickets_current': dist_current['total_tickets'],
        'dist_total_tickets_previous': dist_previous['total_tickets'],
        
        # User permissions
        'user_profile': user_profile,
        'is_admin': user_profile.is_admin,
        'user_locations_count': len(allowed_locations) if not user_profile.is_admin else 'All',
        
        # Product segments
        'bestsellers': product_analysis_current['bestsellers'],
        'least_sellers': product_analysis_current['least_sellers'],
        'slow_movers': product_analysis_current['slow_movers'],
        'rising_stars': product_analysis_current['rising_stars'],
    }
    context['is_admin'] = len(allowed_locations) == 0  # 0 means all locationsx

    
    return render(request, 'dashboard.html', context)

@login_required
def plan_workflow(request):

    try:
        user_profile = request.user.profile
    except:
        return HttpResponseForbidden("Access denied. Contact administrator.")
    
    # Get allowed locations for this user
    allowed_locations_user = user_profile.get_allowed_locations()
    
    # Get filter parameters
    selected_year = request.GET.get('year', '2026')
    selected_start_month = request.GET.get('start_month', '7')
    selected_end_month = request.GET.get('end_month', '7')
    selected_geo = request.GET.get('location', 'all')
    
    # SECURITY CHECK: Validate location access
    if not user_profile.is_admin:
        if selected_geo == 'all':
            if allowed_locations_user:
                selected_geo = allowed_locations_user[0]
            else:
                return HttpResponseForbidden("No locations assigned. Contact administrator.")
        elif selected_geo not in allowed_locations_user:
            messages.warning(request, f"Access denied to location: {selected_geo}")
            selected_geo = allowed_locations_user[0] if allowed_locations_user else 'all'

    aggregation = request.GET.get('aggregation', 'daily')
    show_prev_year = request.GET.get('show_prev_year', 'false')
    
    # Convert to dates
    year = int(selected_year)
    start_month = int(selected_start_month)
    end_month = int(selected_end_month)
    
    start_date = date(year, start_month, 1)
    _, last_day = calendar.monthrange(year, end_month)
    end_date = date(year, end_month, last_day)
    
    # Previous year dates
    prev_year = year - 1
    start_date_py = date(prev_year, start_month, 1)
    end_date_py = date(prev_year, end_date.month, end_date.day)
    
    # Read Excel file
    path = os.path.join(settings.BASE_DIR, 'sales_app', 'data', 'Full Plan workflow.xlsx')
    
    try:
        df = pd.read_excel(path, engine='openpyxl', sheet_name='Main')
        
        print("Excel Columns:", df.columns.tolist())
        print("Sample data:")
        print(df[['location', 'geo', 'Year', 'Month', 'Plan_turnover', 'Plan_tickets', 'Plan_basket']].head(5))
        print(f"\nYear range in Excel: {df['Year'].min()} - {df['Year'].max()}")
        print(f"Month range in Excel: {df['Month'].min()} - {df['Month'].max()}")
        
        df['Year'] = df['Year'].astype(int)
        df['Month'] = df['Month'].astype(int)
        df['plan_date'] = pd.to_datetime(df[['Year', 'Month']].assign(day=1))
        
        # ===== PROCESS CURRENT YEAR DATA =====
        df_current = df.copy()
        start_month_date = pd.Timestamp(start_date.replace(day=1))
        end_month_date = pd.Timestamp(end_date.replace(day=1))
        df_current = df_current[(df_current['plan_date'] >= start_month_date) & (df_current['plan_date'] <= end_month_date)]

        all_geos = sorted(df_current['geo'].unique().tolist())

        if selected_geo != 'all':
            df_current = df_current[df_current['geo'] == selected_geo]
        
        print(f"\nFiltered current year to {len(df_current)} plan records between {start_month_date.strftime('%Y-%m')} and {end_month_date.strftime('%Y-%m')}")
        
        # ===== PROCESS PREVIOUS YEAR DATA =====
        df_prev = df.copy()
        start_month_date_py = pd.Timestamp(start_date_py.replace(day=1))
        end_month_date_py = pd.Timestamp(end_date_py.replace(day=1))
        df_prev = df_prev[(df_prev['plan_date'] >= start_month_date_py) & (df_prev['plan_date'] <= end_month_date_py)]
        
        if selected_geo != 'all':
            df_prev = df_prev[df_prev['geo'] == selected_geo]
        
        print(f"Filtered previous year to {len(df_prev)} plan records between {start_month_date_py.strftime('%Y-%m')} and {end_month_date_py.strftime('%Y-%m')}")
        
        # ===== GET ACTUAL SALES DATA - CURRENT YEAR =====
        actual_query = Sales.objects.filter(
            cd__gte=start_date,
            cd__lte=end_date
        ).exclude(un__in=["მთავარი საწყობი 2", "სატესტო"])
        
        if selected_geo != 'all':
            actual_query = actual_query.filter(un=selected_geo)
        
        daily_actual = list(actual_query.values('un', 'cd').annotate(
            actual_turnover=Sum('tanxa'),
            tickets=Count('zedd', distinct=True)
        ).order_by('cd'))
        
        print(f"\nRetrieved {len(daily_actual)} daily actual records from DB (current year)")
        
        # ===== GET ACTUAL SALES DATA - PREVIOUS YEAR =====
        max_actual_date = max([r['cd'] for r in daily_actual], default=date.today() - timedelta(days=1))
        if hasattr(max_actual_date, 'date'):
            max_actual_date = max_actual_date.date()

        end_date_py_actual = date(prev_year, max_actual_date.month, max_actual_date.day)

        actual_query_py = Sales.objects.filter(
            cd__gte=start_date_py,
            cd__lte=end_date_py_actual
        ).exclude(un__in=["მთავარი საწყობი 2", "სატესტო"])
        
        if selected_geo != 'all':
            actual_query_py = actual_query_py.filter(un=selected_geo)
        
        daily_actual_py = list(actual_query_py.values('un', 'cd').annotate(
            actual_turnover=Sum('tanxa'),
            tickets=Count('zedd', distinct=True)
        ).order_by('cd'))
        
        print(f"Retrieved {len(daily_actual_py)} daily actual records from DB (previous year)")
        
        # ===== EXPAND PLANS TO DAILY =====
        def expand_to_daily(df_source, target_start, target_end):
            daily_records = []
            for _, row in df_source.iterrows():
                geo = row['geo']
                year_row = int(row['Year'])
                month_row = int(row['Month'])
                monthly_plan = float(row['Plan_turnover'])
                monthly_tickets = float(row['Plan_tickets'])
                avg_basket = float(row['Plan_basket'])
                
                days_in_month = calendar.monthrange(year_row, month_row)[1]
                daily_plan_value = monthly_plan / days_in_month
                daily_tickets_value = monthly_tickets / days_in_month
                
                for day in range(1, days_in_month + 1):
                    current_date = date(year_row, month_row, day)
                    if target_start <= current_date <= target_end:
                        daily_records.append({
                            'geo': geo,
                            'date': current_date,
                            'daily_plan': daily_plan_value,
                            'daily_tickets': daily_tickets_value,
                            'avg_basket': avg_basket,
                            'year': year_row,
                            'month': month_row,
                            'day': day
                        })
            return daily_records
        
        plan_daily_records = expand_to_daily(df_current, start_date, end_date)
        plan_daily_records_py = expand_to_daily(df_prev, start_date_py, end_date_py)
        
        print(f"\nExpanded to {len(plan_daily_records)} daily plan records (current year)")
        print(f"Expanded to {len(plan_daily_records_py)} daily plan records (previous year)")
        
        # ===== AGGREGATION HELPER =====
        def aggregate_data(plan_records, actual_records, agg_type, date_range_start, date_range_end):
            labels = []
            plan_values = []
            plan_85_values = []
            actual_values = []
            tickets_plan_values = []
            tickets_actual_values = []
            basket_plan_values = []
            basket_actual_values = []
            
            if agg_type == 'monthly':
                month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                plan_monthly = {}
                tickets_monthly = {}
                basket_monthly = {}
                basket_count = {}
                
                for record in plan_records:
                    month_key = f"{record['year']}-{record['month']:02d}"
                    plan_monthly[month_key] = plan_monthly.get(month_key, 0) + record['daily_plan']
                    tickets_monthly[month_key] = tickets_monthly.get(month_key, 0) + record['daily_tickets']
                    basket_monthly[month_key] = basket_monthly.get(month_key, 0) + record['avg_basket']
                    basket_count[month_key] = basket_count.get(month_key, 0) + 1
                
                actual_monthly = {}
                tickets_actual_monthly = {}
                for record in actual_records:
                    month_key = f"{record['cd'].year}-{record['cd'].month:02d}"
                    actual_monthly[month_key] = actual_monthly.get(month_key, 0) + float(record['actual_turnover'] or 0)
                    tickets_actual_monthly[month_key] = tickets_actual_monthly.get(month_key, 0) + int(record['tickets'] or 0)
                
                current = date_range_start.replace(day=1)
                while current <= date_range_end:
                    month_key = f"{current.year}-{current.month:02d}"
                    labels.append(f"{month_names[current.month-1]} '{str(current.year)[-2:]}")
                    plan_val = plan_monthly.get(month_key, 0)
                    plan_values.append(plan_val)
                    plan_85_values.append(plan_val * 0.85)
                    actual_values.append(actual_monthly.get(month_key, 0))
                    tickets_plan_values.append(tickets_monthly.get(month_key, 0))
                    tickets_actual_values.append(tickets_actual_monthly.get(month_key, 0))
                    if basket_count.get(month_key, 0) > 0:
                        basket_plan_values.append(basket_monthly.get(month_key, 0) / basket_count[month_key])
                    else:
                        basket_plan_values.append(0)
                    actual_rev = actual_monthly.get(month_key, 0)
                    actual_tick = tickets_actual_monthly.get(month_key, 0)
                    basket_actual_values.append(actual_rev / actual_tick if actual_tick > 0 else 0)
                    if current.month == 12:
                        current = current.replace(year=current.year + 1, month=1)
                    else:
                        current = current.replace(month=current.month + 1)
            
            elif agg_type == 'weekly':
                plan_weekly = {}
                tickets_weekly = {}
                basket_weekly = {}
                basket_count = {}
                for record in plan_records:
                    iso_cal = record['date'].isocalendar()
                    week_key = f"{iso_cal[0]}-W{iso_cal[1]:02d}"
                    plan_weekly[week_key] = plan_weekly.get(week_key, 0) + record['daily_plan']
                    tickets_weekly[week_key] = tickets_weekly.get(week_key, 0) + record['daily_tickets']
                    basket_weekly[week_key] = basket_weekly.get(week_key, 0) + record['avg_basket']
                    basket_count[week_key] = basket_count.get(week_key, 0) + 1
                actual_weekly = {}
                tickets_actual_weekly = {}
                for record in actual_records:
                    iso_cal = record['cd'].isocalendar()
                    week_key = f"{iso_cal[0]}-W{iso_cal[1]:02d}"
                    actual_weekly[week_key] = actual_weekly.get(week_key, 0) + float(record['actual_turnover'] or 0)
                    tickets_actual_weekly[week_key] = tickets_actual_weekly.get(week_key, 0) + int(record['tickets'] or 0)
                current = date_range_start
                seen_weeks = set()
                while current <= date_range_end:
                    iso_cal = current.isocalendar()
                    week_key = f"{iso_cal[0]}-W{iso_cal[1]:02d}"
                    if week_key not in seen_weeks:
                        seen_weeks.add(week_key)
                        labels.append(f"W{iso_cal[1]} '{str(iso_cal[0])[-2:]}")
                        plan_val = plan_weekly.get(week_key, 0)
                        plan_values.append(plan_val)
                        plan_85_values.append(plan_val * 0.85)
                        actual_values.append(actual_weekly.get(week_key, 0))
                        tickets_plan_values.append(tickets_weekly.get(week_key, 0))
                        tickets_actual_values.append(tickets_actual_weekly.get(week_key, 0))
                        if basket_count.get(week_key, 0) > 0:
                            basket_plan_values.append(basket_weekly.get(week_key, 0) / basket_count[week_key])
                        else:
                            basket_plan_values.append(0)
                        actual_rev = actual_weekly.get(week_key, 0)
                        actual_tick = tickets_actual_weekly.get(week_key, 0)
                        basket_actual_values.append(actual_rev / actual_tick if actual_tick > 0 else 0)
                    current += timedelta(days=1)
            
            else:  # Daily
                plan_map = {}
                tickets_map = {}
                basket_map = {}
                basket_count = {}
                for record in plan_records:
                    date_key = record['date'].strftime('%Y-%m-%d')
                    plan_map[date_key] = plan_map.get(date_key, 0) + record['daily_plan']
                    tickets_map[date_key] = tickets_map.get(date_key, 0) + record['daily_tickets']
                    basket_map[date_key] = basket_map.get(date_key, 0) + record['avg_basket']
                    basket_count[date_key] = basket_count.get(date_key, 0) + 1
                actual_map = {}
                tickets_actual_map = {}
                for record in actual_records:
                    date_key = record['cd'].strftime('%Y-%m-%d')
                    actual_map[date_key] = actual_map.get(date_key, 0) + float(record['actual_turnover'] or 0)
                    tickets_actual_map[date_key] = tickets_actual_map.get(date_key, 0) + int(record['tickets'] or 0)
                current = date_range_start
                while current <= date_range_end:
                    date_key = current.strftime('%Y-%m-%d')
                    labels.append(current.strftime('%b %d'))
                    plan_val = plan_map.get(date_key, 0)
                    plan_values.append(plan_val)
                    plan_85_values.append(plan_val * 0.85)
                    actual_values.append(actual_map.get(date_key, 0))
                    tickets_plan_values.append(tickets_map.get(date_key, 0))
                    tickets_actual_values.append(tickets_actual_map.get(date_key, 0))
                    if basket_count.get(date_key, 0) > 0:
                        basket_plan_values.append(basket_map.get(date_key, 0) / basket_count[date_key])
                    else:
                        basket_plan_values.append(0)
                    actual_rev = actual_map.get(date_key, 0)
                    actual_tick = tickets_actual_map.get(date_key, 0)
                    basket_actual_values.append(actual_rev / actual_tick if actual_tick > 0 else 0)
                    current += timedelta(days=1)
            
            return {
                'labels': labels,
                'plan_values': plan_values,
                'plan_85_values': plan_85_values,
                'actual_values': actual_values,
                'tickets_plan_values': tickets_plan_values,
                'tickets_actual_values': tickets_actual_values,
                'basket_plan_values': basket_plan_values,
                'basket_actual_values': basket_actual_values
            }
        
        current_data = aggregate_data(plan_daily_records, daily_actual, aggregation, start_date, end_date)
        prev_data = aggregate_data(plan_daily_records_py, daily_actual_py, aggregation, start_date_py, end_date_py)

        # ===== DAYS LEFT LOGIC =====
        # Only meaningful if today falls within the selected year/month range.
        # Otherwise (past or future period) everything days-left related is 0.
        today = date.today()
        if today.year == year and start_month <= today.month <= end_month:
            total_days_in_month = calendar.monthrange(today.year, today.month)[1]
            days_elapsed = today.day
            days_left = total_days_in_month - days_elapsed + 1
        else:
            total_days_in_month = calendar.monthrange(year, end_month)[1]
            days_left = 0

        # ===== CALCULATE KPIs - REVENUE =====
        total_plan = sum(current_data['plan_values'])
        total_plan_85 = sum(current_data['plan_85_values'])
        total_actual = sum(current_data['actual_values'])

        # daily_left_85 = (total_actual - total_plan_85) / days_left if days_left > 0 else 0
        # daily_left_100 = (total_actual - total_plan) / days_left if days_left > 0 else 0

        if (total_actual - total_plan_85) <= 0 and days_left > 0:
            daily_left_85 = (total_actual - total_plan_85) / days_left
        else:          
            daily_left_85 = 0


        if (total_actual - total_plan) <= 0 and days_left > 0:
            daily_left_100 = (total_actual - total_plan) / days_left
        else:          
            daily_left_100 = 0

        plan_achievement = (total_actual / total_plan * 100) if total_plan > 0 else 0
        variance = total_actual - total_plan
        variance_85 = total_actual - total_plan_85
        variance_pct = ((variance / total_plan) * 100) if total_plan > 0 else 0
        variance_pct_85 = ((variance_85 / total_plan_85) * 100) if total_plan_85 > 0 else 0
        
        # ===== CALCULATE KPIs - TICKETS =====
        total_tickets_plan = sum(current_data['tickets_plan_values'])
        total_tickets_actual = sum(current_data['tickets_actual_values'])
        tickets_achievement = (total_tickets_actual / total_tickets_plan * 100) if total_tickets_plan > 0 else 0
        tickets_variance = total_tickets_actual - total_tickets_plan
        tickets_variance_pct = ((tickets_variance / total_tickets_plan) * 100) if total_tickets_plan > 0 else 0
        
        # ===== CALCULATE KPIs - BASKET =====
        avg_basket_plan = sum(current_data['basket_plan_values']) / len(current_data['basket_plan_values']) if len(current_data['basket_plan_values']) > 0 else 0
        avg_basket_actual = sum(current_data['basket_actual_values']) / len(current_data['basket_actual_values']) if len(current_data['basket_actual_values']) > 0 else 0
        basket_achievement = (avg_basket_actual / avg_basket_plan * 100) if avg_basket_plan > 0 else 0
        basket_variance = avg_basket_actual - avg_basket_plan
        basket_variance_pct = ((basket_variance / avg_basket_plan) * 100) if avg_basket_plan > 0 else 0
        
        # ===== LOCATION PERFORMANCE TABLE - REVENUE =====
        location_performance = []
        
        if selected_geo == 'all':
            unique_geos = df_current['geo'].unique()

            for geo in unique_geos:
                # Current Year
                loc_plan_records = [p for p in plan_daily_records if p['geo'] == geo]
                loc_plan = sum([p['daily_plan'] for p in loc_plan_records])
                loc_actual_data = actual_query.filter(un=geo).aggregate(total=Sum('tanxa'))
                loc_actual = float(loc_actual_data['total'] or 0)
                
                # Previous Year
                loc_plan_records_py = [p for p in plan_daily_records_py if p['geo'] == geo]
                loc_plan_py = sum([p['daily_plan'] for p in loc_plan_records_py])
                loc_actual_data_py = actual_query_py.filter(un=geo).aggregate(total=Sum('tanxa'))
                loc_actual_py = float(loc_actual_data_py['total'] or 0)

                # Days left per location — same rule: 0 if not current period
                if today.year == year and start_month <= today.month <= end_month:
                    loc_days_in_month = calendar.monthrange(today.year, today.month)[1]
                    loc_days_left = loc_days_in_month - today.day + 1
                    daily_left_100_per_loc = (loc_actual - loc_plan) / loc_days_left
                    daily_left_85_per_loc = (loc_actual - loc_plan * 0.85) / loc_days_left
                    daily_left_100_per_loc_base = (loc_actual - loc_plan) / loc_days_in_month
                    daily_left_85_per_loc_base = (loc_actual - loc_plan * 0.85) / loc_days_in_month
                else:
                    daily_left_100_per_loc = 0
                    daily_left_85_per_loc = 0
                    daily_left_100_per_loc_base = 0
                    daily_left_85_per_loc_base = 0

                loc_variance = loc_actual - loc_plan
                loc_achievement = (loc_actual / loc_plan * 100) if loc_plan > 0 else 0
                py_loc_achievement = (loc_actual_py / loc_plan_py * 100) if loc_plan_py > 0 else 0
                yoy_growth = ((loc_actual - loc_actual_py) / loc_actual_py * 100) if loc_actual_py > 0 else 0
                yoy_growth_plan = ((loc_plan - loc_plan_py) / loc_plan_py * 100) if loc_plan_py > 0 else 0
                
                location_performance.append({
                    'geo': geo,
                    'plan': loc_plan,
                    'actual': loc_actual,
                    'plan_py': loc_plan_py,
                    'actual_py': loc_actual_py,
                    'variance': loc_variance,
                    'achievement': loc_achievement,
                    'py_loc_achievement': py_loc_achievement,
                    'yoy_growth': yoy_growth,
                    'yoy_growth_plan': yoy_growth_plan,
                    'daily_left_100_per_loc': daily_left_100_per_loc,
                    'daily_left_85_per_loc': daily_left_85_per_loc,
                    'daily_left_100_per_loc_base': daily_left_100_per_loc_base,
                    'daily_left_85_per_loc_base': daily_left_85_per_loc_base
                })
            
            location_performance.sort(key=lambda x: x['achievement'], reverse=True)
        
        # ===== LOCATION PERFORMANCE TABLE - TICKETS =====
        tickets_location_performance = []
        
        if selected_geo == 'all':
            for geo in unique_geos:
                loc_tickets_records = [p for p in plan_daily_records if p['geo'] == geo]
                loc_tickets_plan = sum([p['daily_tickets'] for p in loc_tickets_records])
                loc_tickets_data = actual_query.filter(un=geo).aggregate(total=Count('zedd', distinct=True))
                loc_tickets_actual = int(loc_tickets_data['total'] or 0)
                loc_tickets_records_py = [p for p in plan_daily_records_py if p['geo'] == geo]
                loc_tickets_plan_py = sum([p['daily_tickets'] for p in loc_tickets_records_py])
                loc_tickets_data_py = actual_query_py.filter(un=geo).aggregate(total=Count('zedd', distinct=True))
                loc_tickets_actual_py = int(loc_tickets_data_py['total'] or 0)
                loc_variance = loc_tickets_actual - loc_tickets_plan
                loc_achievement = (loc_tickets_actual / loc_tickets_plan * 100) if loc_tickets_plan > 0 else 0
                yoy_growth = ((loc_tickets_actual - loc_tickets_actual_py) / loc_tickets_actual_py * 100) if loc_tickets_actual_py > 0 else 0
                tickets_location_performance.append({
                    'geo': geo,
                    'plan': loc_tickets_plan,
                    'actual': loc_tickets_actual,
                    'plan_py': loc_tickets_plan_py,
                    'actual_py': loc_tickets_actual_py,
                    'variance': loc_variance,
                    'achievement': loc_achievement,
                    'yoy_growth': yoy_growth
                })
            tickets_location_performance.sort(key=lambda x: x['achievement'], reverse=True)
        
        # ===== LOCATION PERFORMANCE TABLE - BASKET =====
        basket_location_performance = []
        
        if selected_geo == 'all':
            for geo in unique_geos:
                loc_basket_records = [p for p in plan_daily_records if p['geo'] == geo]
                loc_basket_plan = sum([p['avg_basket'] for p in loc_basket_records]) / len(loc_basket_records) if len(loc_basket_records) > 0 else 0
                loc_data = actual_query.filter(un=geo).aggregate(
                    total_rev=Sum('tanxa'),
                    total_tickets=Count('zedd', distinct=True)
                )
                loc_basket_actual = (float(loc_data['total_rev'] or 0) / int(loc_data['total_tickets'] or 1)) if loc_data['total_tickets'] else 0
                loc_basket_records_py = [p for p in plan_daily_records_py if p['geo'] == geo]
                loc_basket_plan_py = sum([p['avg_basket'] for p in loc_basket_records_py]) / len(loc_basket_records_py) if len(loc_basket_records_py) > 0 else 0
                loc_data_py = actual_query_py.filter(un=geo).aggregate(
                    total_rev=Sum('tanxa'),
                    total_tickets=Count('zedd', distinct=True)
                )
                loc_basket_actual_py = (float(loc_data_py['total_rev'] or 0) / int(loc_data_py['total_tickets'] or 1)) if loc_data_py['total_tickets'] else 0
                loc_variance = loc_basket_actual - loc_basket_plan
                loc_achievement = (loc_basket_actual / loc_basket_plan * 100) if loc_basket_plan > 0 else 0
                yoy_change = loc_basket_actual - loc_basket_actual_py
                basket_location_performance.append({
                    'geo': geo,
                    'plan': loc_basket_plan,
                    'actual': loc_basket_actual,
                    'plan_py': loc_basket_plan_py,
                    'actual_py': loc_basket_actual_py,
                    'variance': loc_variance,
                    'achievement': loc_achievement,
                    'yoy_change': yoy_change
                })
            basket_location_performance.sort(key=lambda x: x['achievement'], reverse=True)
        
        # ===== GET LOCATIONS FOR DROPDOWN =====
        if user_profile.is_admin:
            all_geos = sorted(df[df['plan_date'].between(start_month_date_py, end_month_date)]['geo'].unique().tolist())
        else:
            all_geos = allowed_locations_user
        
        # ===== MONTH OPTIONS FOR DROPDOWN =====
        month_options = [
            {'value': '1', 'label': 'January'},
            {'value': '2', 'label': 'February'},
            {'value': '3', 'label': 'March'},
            {'value': '4', 'label': 'April'},
            {'value': '5', 'label': 'May'},
            {'value': '6', 'label': 'June'},
            {'value': '7', 'label': 'July'},
            {'value': '8', 'label': 'August'},
            {'value': '9', 'label': 'September'},
            {'value': '10', 'label': 'October'},
            {'value': '11', 'label': 'November'},
            {'value': '12', 'label': 'December'},
        ]
        
        # ===== EXCEL SUMMARY =====
        excel_summary = df_current.groupby('geo').agg({
            'Plan_turnover': 'sum'
        }).reset_index().sort_values('Plan_turnover', ascending=False).head(10)
        
        excel_data = [
            {'geo': row['geo'], 'Plan_turnover': row['Plan_turnover']}
            for _, row in excel_summary.iterrows()
        ]
        
        file_status = f"✓ Loaded {len(all_geos)} locations with plans from {start_date.strftime('%b %Y')} to {end_date.strftime('%b %Y')}"
        
    except FileNotFoundError:
        current_data = {
            'labels': [], 'plan_values': [], 'plan_85_values': [], 'actual_values': [],
            'tickets_plan_values': [], 'tickets_actual_values': [],
            'basket_plan_values': [], 'basket_actual_values': []
        }
        prev_data = {
            'labels': [], 'plan_values': [], 'plan_85_values': [], 'actual_values': [],
            'tickets_plan_values': [], 'tickets_actual_values': [],
            'basket_plan_values': [], 'basket_actual_values': []
        }
        excel_data = []
        location_performance = []
        tickets_location_performance = []
        basket_location_performance = []
        all_geos = []
        month_options = []
        days_left = 0
        total_plan = total_plan_85 = total_actual = plan_achievement = variance = variance_pct = variance_85 = variance_pct_85 = 0
        daily_left_85 = daily_left_100 = 0
        total_tickets_plan = total_tickets_actual = tickets_achievement = tickets_variance = tickets_variance_pct = 0
        avg_basket_plan = avg_basket_actual = basket_achievement = basket_variance = basket_variance_pct = 0
        file_status = f"✗ Excel file not found at: {path}"
        
    except Exception as e:
        current_data = {
            'labels': [], 'plan_values': [], 'plan_85_values': [], 'actual_values': [],
            'tickets_plan_values': [], 'tickets_actual_values': [],
            'basket_plan_values': [], 'basket_actual_values': []
        }
        prev_data = {
            'labels': [], 'plan_values': [], 'plan_85_values': [], 'actual_values': [],
            'tickets_plan_values': [], 'tickets_actual_values': [],
            'basket_plan_values': [], 'basket_actual_values': []
        }
        excel_data = []
        location_performance = []
        tickets_location_performance = []
        basket_location_performance = []
        all_geos = []
        month_options = []
        days_left = 0
        total_plan = total_plan_85 = total_actual = plan_achievement = variance = variance_pct = variance_85 = variance_pct_85 = 0
        daily_left_85 = daily_left_100 = 0
        total_tickets_plan = total_tickets_actual = tickets_achievement = tickets_variance = tickets_variance_pct = 0
        avg_basket_plan = avg_basket_actual = basket_achievement = basket_variance = basket_variance_pct = 0
        file_status = f"✗ Error: {str(e)}"
        print(f"Error in plan_workflow: {e}")
        import traceback
        traceback.print_exc()

    # ===== CONTEXT =====
    context = {
        'labels': json.dumps(current_data['labels']),

        # Revenue - Current Year
        'plan_values': json.dumps(current_data['plan_values']),
        'plan_85_values': json.dumps(current_data['plan_85_values']),
        'actual_values': json.dumps(current_data['actual_values']),

        # Revenue - Previous Year
        'plan_values_py': json.dumps(prev_data['plan_values']),
        'actual_values_py': json.dumps(prev_data['actual_values']),

        # Tickets - Current Year
        'tickets_plan_values': json.dumps(current_data['tickets_plan_values']),
        'tickets_actual_values': json.dumps(current_data['tickets_actual_values']),

        # Tickets - Previous Year
        'tickets_plan_values_py': json.dumps(prev_data['tickets_plan_values']),
        'tickets_actual_values_py': json.dumps(prev_data['tickets_actual_values']),

        # Basket - Current Year
        'basket_plan_values': json.dumps(current_data['basket_plan_values']),
        'basket_actual_values': json.dumps(current_data['basket_actual_values']),

        # Basket - Previous Year
        'basket_plan_values_py': json.dumps(prev_data['basket_plan_values']),
        'basket_actual_values_py': json.dumps(prev_data['basket_actual_values']),

        # Revenue KPIs
        'total_plan': f"₾{total_plan:,.0f}",
        'total_plan_85': f"₾{total_plan_85:,.0f}",
        'daily_left_85': f"₾{daily_left_85:,.0f}",
        'daily_left_100': f"₾{daily_left_100:,.0f}",
        'total_actual': f"₾{total_actual:,.0f}",
        'plan_achievement': f"{plan_achievement:.1f}",
        'variance': f"₾{variance:,.0f}",
        'variance_85': f"₾{variance_85:,.0f}",
        'variance_pct': f"{variance_pct:+.1f}",
        'variance_pct_85': f"{variance_pct_85:+.1f}",

        # Tickets KPIs
        'total_tickets_plan': f"{total_tickets_plan:,.0f}",
        'total_tickets_actual': f"{total_tickets_actual:,.0f}",
        'tickets_achievement': f"{tickets_achievement:.1f}",
        'tickets_variance': f"{tickets_variance:+,.0f}",
        'tickets_variance_pct': f"{tickets_variance_pct:+.1f}",

        # Basket KPIs
        'avg_basket_plan': f"{avg_basket_plan:.2f}",
        'avg_basket_actual': f"{avg_basket_actual:.2f}",
        'basket_achievement': f"{basket_achievement:.1f}",
        'basket_variance': f"{basket_variance:+.2f}",
        'basket_variance_pct': f"{basket_variance_pct:+.1f}",

        # Location Performance
        'location_performance': location_performance,
        'tickets_location_performance': tickets_location_performance,
        'basket_location_performance': basket_location_performance,

        # Other
        'excel_df': excel_data,
        'all_geos': all_geos,
        'selected_geo': selected_geo,
        'selected_year': selected_year,
        'selected_start_month': selected_start_month,
        'selected_end_month': selected_end_month,
        'month_options': month_options,
        'aggregation': aggregation,
        'show_prev_year': show_prev_year,
        'file_status': file_status,
        'user_profile': user_profile,
        'is_admin': user_profile.is_admin,
    }
    
    return render(request, 'another.html', context)

@login_required
def export_location_csv(request):    
    try:
        user_profile = request.user.profile
    except:
        return HttpResponseForbidden("Access denied. Contact administrator.")
    
    allowed_locations_user = user_profile.get_allowed_locations()
    
    comparison_mode = request.GET.get('comparison', '2025-2024')
    if comparison_mode == '2026-2025':
        current_year = 2026
        previous_year = 2025
    elif comparison_mode == '2026-2024':
        current_year = 2026
        previous_year = 2024
    else:
        current_year = 2025
        previous_year = 2024
    
    # Date filters
    start_date_str = request.GET.get('start_date', f'{current_year}-01-01')
    end_date_str = request.GET.get('end_date', f'{current_year}-12-31')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except:
        start_date = date(current_year, 1, 1)
    
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except:
        end_date = date(current_year, 12, 31)
    
    # Location filter - handle multiple selections with SECURITY CHECK
    selected_locations = request.GET.getlist('un_filter')
    
    if not user_profile.is_admin:
        if not selected_locations or 'all' in selected_locations:
            # Non-admin can't export all - restrict to their locations
            selected_locations = allowed_locations_user
        else:
            # Filter out unauthorized locations
            unauthorized = set(selected_locations) - set(allowed_locations_user)
            if unauthorized:
                messages.warning(request, f"Export access denied to: {', '.join(unauthorized)}")
                selected_locations = [loc for loc in selected_locations if loc in allowed_locations_user]
            
            if not selected_locations:
                selected_locations = allowed_locations_user
    
    if not selected_locations and not user_profile.is_admin:
        return HttpResponseForbidden("You don't have access to export any locations.")
    
    # If admin selected 'all', reset to empty list
    if user_profile.is_admin and (not selected_locations or 'all' in request.GET.getlist('un_filter')):
        selected_locations = []
    
    selected_category = request.GET.get('category', 'all')
    selected_product = request.GET.get('prod_filter', 'all')
    selected_campaign = request.GET.get('campaign_filter', 'all')
    
    def get_location_data(year, start_dt, end_dt):
        """Helper function to get data for a specific year"""
        # Create timezone-aware datetimes
        start_datetime = timezone.make_aware(datetime.combine(start_dt, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(end_dt, datetime.max.time()))
        
        # Build the base query
        query = Sales.objects.filter(
            prodt='selling item',
            cd__year=year,
            cd__gte=start_datetime,
            cd__lte=end_datetime
        ).exclude(un__in=["მთავარი საწყობი 2", "სატესტო"]).exclude(tanxa=0)
        
        # Debug: Print what filters are being applied
        print(f"Year: {year}")
        print(f"Locations filter: {selected_locations}")
        print(f"Category filter: {selected_category}")
        print(f"Product filter: {selected_product}")
        print(f"Campaign filter: {selected_campaign}")
        
        # Apply filters one by one
        if selected_locations and len(selected_locations) > 0:
            print(f"Applying location filter: {selected_locations}")
            query = query.filter(un__in=selected_locations)
        
        if selected_category and selected_category != 'all':
            print(f"Applying category filter: {selected_category}")
            query = query.filter(prodg=selected_category)
        
        if selected_product and selected_product != 'all':
            print(f"Applying product filter: {selected_product}")
            query = query.filter(prod=selected_product)
        
        if selected_campaign and selected_campaign != 'all':
            print(f"Applying campaign filter: {selected_campaign}")
            query = query.filter(actions=selected_campaign)
        
        # Debug: Print query count
        print(f"Query count after filters: {query.count()}")
        
        # Create filtered subquery for cross-selling calculations with same filters
        filtered_tickets = Sales.objects.filter(
            prodt='selling item',
            cd__year=year,
            cd__gte=start_datetime,
            cd__lte=end_datetime
        ).exclude(un__in=["მთავარი საწყობი 2", "სატესტო"]).exclude(tanxa=0).exclude(prodg='POP')
        
        if selected_locations and len(selected_locations) > 0:
            filtered_tickets = filtered_tickets.filter(un__in=selected_locations)
        
        if selected_category and selected_category != 'all':
            filtered_tickets = filtered_tickets.filter(prodg=selected_category)
        
        if selected_product and selected_product != 'all':
            filtered_tickets = filtered_tickets.filter(prod=selected_product)
        
        if selected_campaign and selected_campaign != 'all':
            filtered_tickets = filtered_tickets.filter(actions=selected_campaign)
        
        # Get location aggregations
        location_data = query.values('un').annotate(
            total=Sum('tanxa'),
            tickets=Count('zedd', distinct=True),
            quantity=Sum('raod'),
            three_plus=Count(
                'zedd',
                distinct=True,
                filter=Q(
                    zedd__in=filtered_tickets
                        .values('zedd')
                        .annotate(c=Count('idreal1'))
                        .filter(c__gte=3)
                        .values('zedd')
                )
            ),
            one_count=Count(
                'zedd',
                distinct=True,
                filter=Q(
                    zedd__in=filtered_tickets
                        .values('zedd')
                        .annotate(c=Count('idreal1'))
                        .filter(c=1)
                        .values('zedd')
                )
            )
        ).annotate(
            avg_basket=ExpressionWrapper(
                F('total') * 1.0 / F('tickets'),
                output_field=FloatField()
            ),
            three_plus_ratio=ExpressionWrapper(
                (F('three_plus') * 100.0) / F('tickets'),
                output_field=FloatField()
            ),
            one_ratio=ExpressionWrapper(
                (F('one_count') * 100.0) / F('tickets'),
                output_field=FloatField()
            )
        ).order_by('-total')
        
        return location_data
    
    # Get data for both years
    previous_start = start_date.replace(year=previous_year)
    previous_end = end_date.replace(year=previous_year)
    
    current_data = list(get_location_data(current_year, start_date, end_date))
    previous_data = list(get_location_data(previous_year, previous_start, previous_end))
    
    # Create Excel workbook
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    total_font = Font(bold=True, size=11)
    info_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    def create_sheet(ws, data, year, sheet_name):
        """Create a formatted sheet with location data"""
        ws.title = sheet_name
        
        # Add header information
        ws['A1'] = 'Location Performance Report'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A2'] = 'Year:'
        ws['B2'] = year
        ws['A2'].font = info_font
        
        ws['A3'] = 'Period:'
        if year == current_year:
            ws['B3'] = f'{start_date} to {end_date}'
        else:
            ws['B3'] = f'{previous_start} to {previous_end}'
        ws['A3'].font = info_font
        
        ws['A4'] = 'Category:'
        ws['B4'] = selected_category
        ws['A4'].font = info_font
        
        ws['A5'] = 'Product:'
        ws['B5'] = selected_product if selected_product != 'all' else 'All'
        ws['A5'].font = info_font
        
        ws['A6'] = 'Campaign:'
        ws['B6'] = selected_campaign if selected_campaign != 'all' else 'All'
        ws['A6'].font = info_font
        
        # Column headers (row 8)
        headers = [
            'Location', 'Total Amount', 'Tickets', 'Quantity', 
            'Avg Basket', '3+ Items', '1 Item', '3+ Ratio (%)', '1 Item Ratio (%)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row_num = 9
        total_revenue = 0
        total_tickets = 0
        total_quantity = 0
        total_3plus = 0
        total_1item = 0
        
        for row_data in data:
            ws.cell(row=row_num, column=1, value=row_data['un'])
            ws.cell(row=row_num, column=2, value=round(row_data['total'], 2) if row_data['total'] else 0)
            ws.cell(row=row_num, column=3, value=row_data['tickets'])
            ws.cell(row=row_num, column=4, value=row_data['quantity'])
            ws.cell(row=row_num, column=5, value=round(row_data['avg_basket'], 2) if row_data['avg_basket'] else 0)
            ws.cell(row=row_num, column=6, value=row_data['three_plus'])
            ws.cell(row=row_num, column=7, value=row_data['one_count'])
            ws.cell(row=row_num, column=8, value=round(row_data['three_plus_ratio'], 2) if row_data['three_plus_ratio'] else 0)
            ws.cell(row=row_num, column=9, value=round(row_data['one_ratio'], 2) if row_data['one_ratio'] else 0)
            
            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row_num, column=col).border = border
            
            # Number formatting
            ws.cell(row=row_num, column=2).number_format = '#,##0.00'
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            ws.cell(row=row_num, column=8).number_format = '0.00'
            ws.cell(row=row_num, column=9).number_format = '0.00'
            
            # Accumulate totals
            total_revenue += row_data['total'] or 0
            total_tickets += row_data['tickets'] or 0
            total_quantity += row_data['quantity'] or 0
            total_3plus += row_data['three_plus'] or 0
            total_1item += row_data['one_count'] or 0
            
            row_num += 1
        
        # Add totals row
        row_num += 1
        avg_basket_total = total_revenue / total_tickets if total_tickets > 0 else 0
        ratio_3plus = (total_3plus / total_tickets * 100) if total_tickets > 0 else 0
        ratio_1item = (total_1item / total_tickets * 100) if total_tickets > 0 else 0
        
        ws.cell(row=row_num, column=1, value='TOTAL')
        ws.cell(row=row_num, column=2, value=round(total_revenue, 2))
        ws.cell(row=row_num, column=3, value=total_tickets)
        ws.cell(row=row_num, column=4, value=total_quantity)
        ws.cell(row=row_num, column=5, value=round(avg_basket_total, 2))
        ws.cell(row=row_num, column=6, value=total_3plus)
        ws.cell(row=row_num, column=7, value=total_1item)
        ws.cell(row=row_num, column=8, value=round(ratio_3plus, 2))
        ws.cell(row=row_num, column=9, value=round(ratio_1item, 2))
        
        # Style totals row
        for col in range(1, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
        
        ws.cell(row=row_num, column=2).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws.cell(row=row_num, column=8).number_format = '0.00'
        ws.cell(row=row_num, column=9).number_format = '0.00'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Create sheets for current and previous year
    ws_current = wb.active
    create_sheet(ws_current, current_data, current_year, f'{current_year}')
    
    ws_previous = wb.create_sheet(title=f'{previous_year}')
    create_sheet(ws_previous, previous_data, previous_year, f'{previous_year}')
    
    # Create comparison sheet
    ws_comparison = wb.create_sheet(title='Comparison')
    ws_comparison['A1'] = f'{previous_year} vs {current_year} Comparison'
    ws_comparison['A1'].font = Font(bold=True, size=14)
    
    # Comparison headers
    comp_headers = [
        'Location', 
        f'{previous_year} Revenue', f'{current_year} Revenue', 'Revenue Change', 'Revenue Change %',
        f'{previous_year} Tickets', f'{current_year} Tickets', 'Tickets Change', 'Tickets Change %'
    ]
    
    for col, header in enumerate(comp_headers, 1):
        cell = ws_comparison.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Create comparison data
    prev_dict = {row['un']: row for row in previous_data}
    curr_dict = {row['un']: row for row in current_data}
    all_locations = sorted(set(list(prev_dict.keys()) + list(curr_dict.keys())))
    
    row_num = 4
    for location in all_locations:
        prev = prev_dict.get(location, {})
        curr = curr_dict.get(location, {})
        
        prev_revenue = prev.get('total', 0) or 0
        curr_revenue = curr.get('total', 0) or 0
        revenue_change = curr_revenue - prev_revenue
        revenue_change_pct = ((revenue_change / prev_revenue) * 100) if prev_revenue > 0 else 0
        
        prev_tickets = prev.get('tickets', 0) or 0
        curr_tickets = curr.get('tickets', 0) or 0
        tickets_change = curr_tickets - prev_tickets
        tickets_change_pct = ((tickets_change / prev_tickets) * 100) if prev_tickets > 0 else 0
        
        ws_comparison.cell(row=row_num, column=1, value=location)
        ws_comparison.cell(row=row_num, column=2, value=round(prev_revenue, 2))
        ws_comparison.cell(row=row_num, column=3, value=round(curr_revenue, 2))
        ws_comparison.cell(row=row_num, column=4, value=round(revenue_change, 2))
        ws_comparison.cell(row=row_num, column=5, value=round(revenue_change_pct, 2))
        ws_comparison.cell(row=row_num, column=6, value=prev_tickets)
        ws_comparison.cell(row=row_num, column=7, value=curr_tickets)
        ws_comparison.cell(row=row_num, column=8, value=tickets_change)
        ws_comparison.cell(row=row_num, column=9, value=round(tickets_change_pct, 2))
        
        # Apply conditional formatting colors
        for col in [4, 5, 8, 9]:
            cell = ws_comparison.cell(row=row_num, column=col)
            value = cell.value
            if value > 0:
                cell.font = Font(color="10B981")
            elif value < 0:
                cell.font = Font(color="EF4444")
        
        # Apply borders
        for col in range(1, 10):
            ws_comparison.cell(row=row_num, column=col).border = border
        
        row_num += 1
    
    # Adjust comparison sheet column widths
    ws_comparison.column_dimensions['A'].width = 25
    for col in range(2, 10):
        ws_comparison.column_dimensions[get_column_letter(col)].width = 16
    
    # Create filename
    filename_parts = [f'location_report_{current_year}_vs_{previous_year}']
    if selected_locations:
        filename_parts.append(f'{len(selected_locations)}locations')
    if selected_category != 'all':
        filename_parts.append(selected_category.replace(' ', '_'))
    filename_parts.append(f'{start_date.strftime("%Y%m%d")}-{end_date.strftime("%Y%m%d")}')
    
    filename = '_'.join(filename_parts) + '.xlsx'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def query(request):
    try:
        user_profile = request.user.profile
    except:
        return HttpResponseForbidden("Access denied. Contact administrator.")
    
    # ADMIN ONLY for query interface
    if not user_profile.is_admin:
        return HttpResponseForbidden("Only administrators can access the SQL query interface.")
    
    results = None
    columns = None
    query_text = ""
    error_message = None
    
    # Handle Excel export
    if request.method == 'POST' and 'export_excel' in request.POST:
        query_text = request.POST.get('sql_query', '').strip()
        
        if query_text:
            # Basic SQL injection prevention - allow SELECT and WITH (CTE) statements
            query_upper = query_text.upper().strip()
            if not (query_upper.startswith('SELECT') or query_upper.startswith('WITH')):
                error_message = "Only SELECT queries (including CTEs with WITH) are allowed for security reasons."
            # Check for dangerous keywords
            elif any(keyword in query_upper for keyword in ['DROP', 'DELETE', 'INSERT', 'UPDATE', 'ALTER', 'CREATE', 'TRUNCATE', 'EXEC', 'EXECUTE']):
                error_message = "Detected prohibited SQL keywords. Only SELECT queries are allowed."
            else:
                try:
                    with connection.cursor() as cursor:
                        cursor.execute(query_text)
                        results = cursor.fetchall()
                        columns = [col[0] for col in cursor.description]
                    
                    # Create Excel file
                    return export_to_excel(results, columns)
                    
                except Exception as e:
                    error_message = f"Query Error: {str(e)}"
    
    # Handle regular query execution
    elif request.method == 'POST':
        query_text = request.POST.get('sql_query', '').strip()
        
        if query_text:
            # Security validation
            query_upper = query_text.upper().strip()
            if not (query_upper.startswith('SELECT') or query_upper.startswith('WITH')):
                error_message = "Only SELECT queries (including CTEs with WITH) are allowed for security reasons."
            elif any(keyword in query_upper for keyword in ['DROP', 'DELETE', 'INSERT', 'UPDATE', 'ALTER', 'CREATE', 'TRUNCATE', 'EXEC', 'EXECUTE']):
                error_message = "Detected prohibited SQL keywords. Only SELECT queries are allowed."
            else:
                try:
                    with connection.cursor() as cursor:
                        cursor.execute(query_text)
                        results = cursor.fetchall()
                        columns = [col[0] for col in cursor.description]
                    
                    if not results:
                        messages.info(request, "Query executed successfully but returned no results.")
                    else:
                        messages.success(request, f"Query executed successfully! {len(results)} rows returned.")
                        
                except Exception as e:
                    error_message = f"Query Error: {str(e)}"
    
    context = {
        'results': results,
        'columns': columns,
        'query_text': query_text,
        'error_message': error_message,
        'user_profile': user_profile,
        'is_admin': user_profile.is_admin,
    }
    
    return render(request, 'query.html', context)

def export_to_excel(results, columns):
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Query Results"
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    cell_alignment = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Write headers
    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, row_data in enumerate(results, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = cell_alignment
            cell.border = border
            
            # Alternate row colors for better readability
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Auto-adjust column widths
    for col_num in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_num)
        
        # Calculate max length in column
        max_length = len(str(columns[col_num - 1]))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        # Set column width (with some padding)
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=query_results.xlsx'
    
    # Save workbook to response
    wb.save(response)
    
    return response

@login_required
def employee_analytics(request):
    """Optimized employee analytics view with reduced queries and better performance"""
    
    # ==================== SETUP & VALIDATION ====================
    try:
        user_profile = request.user.profile
    except:
        return HttpResponseForbidden("Access denied. Contact administrator.")
    
    allowed_locations = user_profile.get_allowed_locations()
    
    # Year selection
    comparison_mode = request.GET.get('comparison', '2026-2025')
    if comparison_mode == '2026-2025':
        current_year, previous_year = 2026, 2025
    elif comparison_mode == '2026-2024':
        current_year, previous_year = 2026, 2024
    else:
        current_year, previous_year = 2025, 2024
    today = date.today()
    month_current = today.month
    start_date = 1
    last_day = calendar.monthrange(today.year, today.month)[1]
    if len(str(month_current)) > 1:
        month_current = month_current   
    else:
        month_current = '0' + str(month_current)
    # Date parsing
    start_date_str = request.GET.get('start_date', f'{current_year}-{month_current}-{start_date}')
    end_date_str = request.GET.get('end_date', f'{current_year}-{month_current}-{last_day}')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except:
        start_date = date(current_year, 1, 1)
    
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except:
        end_date = date(current_year, 12, 31)
    
    # Location security check
    selected_locations = request.GET.getlist('un_filter')
    
    if not user_profile.is_admin:
        if not selected_locations or 'all' in selected_locations:
            selected_locations = allowed_locations
        else:
            unauthorized = set(selected_locations) - set(allowed_locations)
            if unauthorized:
                messages.warning(request, f"Access denied to: {', '.join(unauthorized)}")
                selected_locations = [loc for loc in selected_locations if loc in allowed_locations]
            if not selected_locations:
                selected_locations = allowed_locations
    
    if not selected_locations and not user_profile.is_admin:
        return HttpResponseForbidden("You don't have access to any locations.")
    
    # Display selected location
    if user_profile.is_admin and (not selected_locations or 'all' in request.GET.getlist('un_filter')):
        selected_un = 'all'
        selected_locations = []
    else:
        selected_un = selected_locations[0] if len(selected_locations) == 1 else 'multiple'
    
    # Other filters
    selected_category = request.GET.get('category', 'all')
    selected_employee = request.GET.get('employee_filter', 'all')
    selected_products = request.GET.getlist('prod_filter')
    
    # Adjust dates to current year
    start_date = start_date.replace(year=current_year)
    end_date = end_date.replace(year=current_year)
    
    # Previous year dates
    previous_start = start_date.replace(year=previous_year)
    previous_end = end_date.replace(year=previous_year)
    
    # Timezone-aware datetimes
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    previous_start_datetime = timezone.make_aware(datetime.combine(previous_start, datetime.min.time()))
    previous_end_datetime = timezone.make_aware(datetime.combine(previous_end, datetime.max.time()))
    
    date_filter_current = {
        'cd__year': current_year,
        'cd__gte': start_datetime,
        'cd__lte': end_datetime
    }
    
    date_filter_previous = {
        'cd__year': previous_year,
        'cd__gte': previous_start_datetime,
        'cd__lte': previous_end_datetime
    }
    
    # ==================== HELPER FUNCTIONS ====================
    
    def apply_filters(q):
        if selected_employee != "all" and not selected_locations:
            return q.none()

        if selected_locations:
            q = q.filter(un__in=selected_locations)

        if selected_category != "all":
            q = q.filter(prodg=selected_category)

        if selected_employee != "all":
            q = q.filter(tanam=selected_employee)
        
        if selected_products:
            q = q.filter(prod__in=selected_products)

        return q
    
    def get_base_queryset(is_current=True):
        if is_current:
            q = Sales.objects.filter(**date_filter_current).exclude(un__in=["მთავარი საწყობი 2", "სატესტო"])
        else:
            q = Sales.objects.filter(**date_filter_previous).exclude(un__in=["მთავარი საწყობი 2", "სატესტო"])
        return apply_filters(q)
    
    # ==================== OPTIMIZED DATA FETCHING ====================
    
    def get_employee_performance_optimized(is_current=True):
        q = get_base_queryset(is_current)
        
        employee_base_stats = q.values('tanam', 'un').annotate(
            total_revenue=Sum('tanxa'),
            total_revenue_skincare_eligible=Sum('tanxa', filter=Q(~Q(prodg='POP'))),
            skincare_turnover=Sum('tanxa', filter=Q(prodg='SKIN CARE')),
            total_tickets=Count('zedd', distinct=True),
            total_items=Count('zedd', filter=Q(~Q(prodg='POP'))),
            discount_given=Sum('discount_price'),
            std_price_total=Sum('std_price')
        ).order_by('-total_revenue')[:20]
        
        employee_list = list(employee_base_stats)
        
        if not employee_list:
            return []
        
        employee_names = [emp['tanam'] for emp in employee_list]

        employee_category_query = (Sales.objects
            .filter(**date_filter_current if is_current else date_filter_previous)
            .filter(tanam__in=employee_names))

        if selected_locations:
            employee_category_query = employee_category_query.filter(un__in=selected_locations)
        if selected_category != 'all':
            employee_category_query = employee_category_query.filter(prodg=selected_category)

        employee_category_data = employee_category_query.values('tanam', 'prodg').annotate(
            total=Sum('tanxa')
        ).order_by('tanam', '-total')

        employee_categories = {}
        for record in employee_category_data:
            emp_name = record['tanam']
            if emp_name not in employee_categories:
                employee_categories[emp_name] = []
            employee_categories[emp_name].append({
                'category': record['prodg'] or 'Unknown',
                'value': float(record['total'] or 0)
            })
        
        cross_sell_query = (Sales.objects
            .filter(**date_filter_current if is_current else date_filter_previous)
            .filter(prodt='selling item', tanam__in=employee_names)
            .exclude(tanxa=0)
            .exclude(prodg='POP'))
        
        if selected_locations:
            cross_sell_query = cross_sell_query.filter(un__in=selected_locations)
        if selected_category != 'all':
            cross_sell_query = cross_sell_query.filter(prodg=selected_category)
        
        cross_sell_data = cross_sell_query.values('tanam', 'zedd').annotate(
            item_count=Count('zedd')
        )
        
        employee_cross_sell = {}
        for record in cross_sell_data:
            emp_name = record['tanam']
            if emp_name not in employee_cross_sell:
                employee_cross_sell[emp_name] = {
                    'total': 0,
                    'one_item': 0,
                    'two_item': 0,
                    'three_plus': 0
                }
            
            employee_cross_sell[emp_name]['total'] += 1
            item_count = record['item_count']
            
            if item_count == 1:
                employee_cross_sell[emp_name]['one_item'] += 1
            elif item_count == 2:
                employee_cross_sell[emp_name]['two_item'] += 1
            elif item_count >= 3:
                employee_cross_sell[emp_name]['three_plus'] += 1
        
        results = []
        for emp in employee_list:
            emp_name = emp['tanam'] or 'Unknown'
            
            total_rev_skincare = emp.get('total_revenue_skincare_eligible') or 0
            skincare_turnover = emp.get('skincare_turnover') or 0
            skincare_percentage = (
                float(skincare_turnover) / float(total_rev_skincare) * 100
                if total_rev_skincare > 0
                else 0
            )
            
            avg_basket = float(emp['total_revenue'] or 0) / emp['total_tickets'] if emp['total_tickets'] > 0 else 0
            items_per_ticket = emp['total_items'] / emp['total_tickets'] if emp['total_tickets'] > 0 else 0
            discount_rate = (1 - (emp['discount_given'] / emp['std_price_total'])) * 100 if emp['std_price_total'] and emp['std_price_total'] > 0 else 0
            
            cs_data = employee_cross_sell.get(emp_name, {
                'total': 0,
                'one_item': 0,
                'two_item': 0,
                'three_plus': 0
            })

            category_breakdown = employee_categories.get(emp_name, [])[:8]
            category_labels = [cat['category'] for cat in category_breakdown]
            category_values = [cat['value'] for cat in category_breakdown]
            
            total_cs_tickets = cs_data['total']
            
            if total_cs_tickets > 0:
                cross_sell_pct = (cs_data['three_plus'] / total_cs_tickets) * 100
                one_item_pct = (cs_data['one_item'] / total_cs_tickets) * 100
                two_item_pct = (cs_data['two_item'] / total_cs_tickets) * 100
                three_plus_pct = cross_sell_pct
                avg_items_per_ticket = items_per_ticket
            else:
                cross_sell_pct = 0
                one_item_pct = 0
                two_item_pct = 0
                three_plus_pct = 0
                avg_items_per_ticket = 0
            
            results.append({
                'name': emp_name,
                'location': emp['un'] or 'Unknown',
                'revenue': float(emp['total_revenue'] or 0),
                'skincare_percentage': skincare_percentage,
                'tickets': emp['total_tickets'],
                'items': emp['total_items'],
                'avg_basket': avg_basket,
                'items_per_ticket': items_per_ticket,
                'discount_rate': discount_rate,
                'cross_sell_pct': cross_sell_pct,
                'one_item_pct': one_item_pct,
                'two_item_pct': two_item_pct,
                'three_plus_pct': three_plus_pct,
                'avg_items_per_ticket': avg_items_per_ticket,
                'cross_sell_tickets': cs_data['three_plus'],
                'one_item_tickets': cs_data['one_item'],
                'two_item_tickets': cs_data['two_item'],
                'three_plus_tickets': cs_data['three_plus'],
                'category_labels_json': json.dumps(category_labels),
                'category_values_json': json.dumps(category_values),
            })
        
        return results
    
    def get_all_category_leaders_optimized(is_current=True):
        top_categories_query = Sales.objects.filter(**date_filter_current if is_current else date_filter_previous)
        if selected_locations:
            top_categories_query = top_categories_query.filter(un__in=selected_locations)
        
        top_categories = list(
            top_categories_query
            .values('prodg')
            .annotate(total=Sum('tanxa'))
            .order_by('-total')[:10]
            .values_list('prodg', flat=True)
        )
        
        if not top_categories:
            return []
        
        q = Sales.objects.filter(**date_filter_current if is_current else date_filter_previous)
        if selected_locations:
            q = q.filter(un__in=selected_locations)
        q = q.filter(prodg__in=top_categories)
        
        all_category_stats = list(
            q.values('prodg', 'tanam')
            .annotate(
                total_revenue=Sum('tanxa'),
                total_tickets=Count('zedd', distinct=True),
                total_items=Count('zedd')
            )
            .order_by('prodg', '-total_revenue')
        )
        
        cross_sell_query = (Sales.objects
            .filter(**date_filter_current if is_current else date_filter_previous)
            .filter(prodt='selling item', prodg__in=top_categories)
            .exclude(tanxa=0)
            .exclude(prodg='POP'))
        
        if selected_locations:
            cross_sell_query = cross_sell_query.filter(un__in=selected_locations)
        
        cross_sell_data = cross_sell_query.values('prodg', 'tanam', 'zedd').annotate(
            item_count=Count('zedd')
        )
        
        category_employee_cs = {}
        for record in cross_sell_data:
            cat = record['prodg']
            emp = record['tanam']
            key = (cat, emp)
            
            if key not in category_employee_cs:
                category_employee_cs[key] = {'total': 0, 'three_plus': 0}
            
            category_employee_cs[key]['total'] += 1
            if record['item_count'] >= 3:
                category_employee_cs[key]['three_plus'] += 1
        
        category_leaders = []
        for category in top_categories:
            cat_stats = [s for s in all_category_stats if s['prodg'] == category]
            cat_stats.sort(key=lambda x: x['total_revenue'], reverse=True)
            top_10 = cat_stats[:10]
            
            performers = []
            for emp in top_10:
                emp_name = emp['tanam'] or 'Unknown'
                key = (category, emp['tanam'])
                cs_data = category_employee_cs.get(key, {'total': 0, 'three_plus': 0})
                cross_sell_pct = (cs_data['three_plus'] / cs_data['total'] * 100) if cs_data['total'] > 0 else 0
                
                performers.append({
                    'name': emp_name,
                    'revenue': float(emp['total_revenue'] or 0),
                    'tickets': emp['total_tickets'],
                    'items': emp['total_items'],
                    'cross_sell_pct': cross_sell_pct
                })
            
            category_leaders.append({
                'category': category,
                'performers_current': performers if is_current else [],
                'performers_previous': [] if is_current else performers
            })
        
        return category_leaders
    
    # ==================== EXECUTE DATA FETCHING ====================
    
    print("Starting employee analytics data fetch...")
    start_time = timezone.now()
    
    employees_current = get_employee_performance_optimized(is_current=True)
    employees_previous = get_employee_performance_optimized(is_current=False)
    
    print(f"Employee performance fetched in {(timezone.now() - start_time).total_seconds():.2f}s")
    
    employees_previous_dict = {emp['name']: emp for emp in employees_previous}
    
    for emp in employees_current:
        prev_data = employees_previous_dict.get(emp['name'], {})
        emp['revenue_previous'] = prev_data.get('revenue', 0)
        emp['tickets_previous'] = prev_data.get('tickets', 0)
        emp['revenue_change'] = ((emp['revenue'] - emp['revenue_previous']) / emp['revenue_previous'] * 100) if emp['revenue_previous'] > 0 else 0
        emp['tickets_change'] = ((emp['tickets'] - emp['tickets_previous']) / emp['tickets_previous'] * 100) if emp['tickets_previous'] > 0 else 0
    
    category_leaders_current = get_all_category_leaders_optimized(is_current=True)
    category_leaders_previous = get_all_category_leaders_optimized(is_current=False)
    
    print(f"Category leaders fetched in {(timezone.now() - start_time).total_seconds():.2f}s")
    
    category_leaders = []
    for cat_current in category_leaders_current:
        cat_name = cat_current['category']
        cat_previous = next(
            (c for c in category_leaders_previous if c['category'] == cat_name),
            {'category': cat_name, 'performers_previous': []}
        )
        category_leaders.append({
            'category': cat_name,
            'performers_current': cat_current['performers_current'],
            'performers_previous': cat_previous['performers_previous']
        })
    
    top_categories = [cat['category'] for cat in category_leaders]
    
    print(f"All category data processed in {(timezone.now() - start_time).total_seconds():.2f}s")
    
    # ==================== EMPLOYEE INSIGHTS (no extra queries) ====================
    # Derived from employees_current in Python — zero additional DB hits
    
    def safe_max(lst, key):
        """Return the employee with the highest value for key, or None if empty."""
        filtered = [e for e in lst if e.get(key) is not None]
        return max(filtered, key=lambda x: x[key]) if filtered else None

    employee_insights = {}
    if employees_current:
        employee_insights = {
            'revenue_leader':       safe_max(employees_current, 'revenue'),
            'best_cross_seller':    safe_max(employees_current, 'cross_sell_pct'),
            'highest_avg_basket':   safe_max(employees_current, 'avg_basket'),
            'most_tickets':         safe_max(employees_current, 'tickets'),
            'skincare_specialist':  safe_max(employees_current, 'skincare_percentage'),
            'top_yoy_growth':       safe_max(employees_current, 'revenue_change'),
            'most_items_per_ticket': safe_max(employees_current, 'avg_items_per_ticket'),
        }

    # ==================== GET FILTER OPTIONS ====================
    
    if user_profile.is_admin:
        all_locations = list(
            Sales.objects
            .filter(cd__year=current_year)
            .values_list('un', flat=True)
            .distinct()
            .order_by('un')
        )
    else:
        all_locations = allowed_locations
    
    filter_base_query = Sales.objects.filter(cd__year=current_year)
    
    if selected_locations:
        filter_base_query = filter_base_query.filter(un__in=selected_locations)
    
    all_categories = list(
        filter_base_query
        .values_list('prodg', flat=True)
        .distinct()
        .order_by('prodg')
    )
    
    all_employees = list(
        filter_base_query
        .values_list('tanam', flat=True)
        .distinct()
        .order_by('tanam')
    )
    
    all_products = list(
        filter_base_query
        .values_list('prod', flat=True)
        .distinct()
        .order_by('prod')
    )
    
    date_range_text = f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d')}, {current_year}"
    
    total_time = (timezone.now() - start_time).total_seconds()
    print(f"Total employee analytics load time: {total_time:.2f}s")
    
    # ==================== BUILD CONTEXT ====================
    
    context = {
        'comparison_mode': comparison_mode,
        'current_year': current_year,
        'previous_year': previous_year,
        'date_range_text': date_range_text,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        
        'employees_current': employees_current,
        'employee_insights': employee_insights,   # NEW — used by insight cards
        'category_leaders': category_leaders,
        'top_categories': top_categories,
        
        'all_locations': all_locations,
        'all_categories': all_categories,
        'all_employees': all_employees,
        'all_products': all_products,
        'selected_un': selected_un,
        'selected_locations': selected_locations,
        'selected_category': selected_category,
        'selected_employee': selected_employee,
        'selected_products': selected_products,
        
        'user_profile': user_profile,
        'is_admin': user_profile.is_admin,
        'user_locations_count': len(allowed_locations) if not user_profile.is_admin else 0,
        
        'load_time': f"{total_time:.2f}s",
    }
    
    return render(request, 'employee_analytics.html', context)


@login_required
def insights(request):
    """
    Generate AI-powered insights by comparing current selection with up to 2 previous years
    """
    
    # Get filter parameters - same as dashboard
    try:
        user_profile = request.user.profile
    except:
        return HttpResponseForbidden("Access denied. Contact administrator.")
    
    # Get allowed locations for this user
    allowed_locations = user_profile.get_allowed_locations()
    
    # Get filter parameters - same as dashboard
    comparison_mode = request.GET.get('comparison', '2025-2024')
    if comparison_mode == '2026-2025':
        current_year = 2026
        previous_year = 2025
        two_years_ago = 2024
    elif comparison_mode == '2026-2024':
        current_year = 2026
        previous_year = 2024
        two_years_ago = None  # Not enough data
    else:
        current_year = 2025
        previous_year = 2024
        two_years_ago = 2023

    # Get date range
    start_date_str = request.GET.get('start_date', f'{current_year}-01-01')
    end_date_str = request.GET.get('end_date', f'{current_year}-12-31')
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except:
        start_date = date(current_year, 1, 1)
    
    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except:
        end_date = date(current_year, 12, 31)
    
    # Handle location filtering with SECURITY CHECK
    selected_locations = request.GET.getlist('un_filter')
    
    if not user_profile.is_admin:
        if not selected_locations or 'all' in selected_locations:
            selected_locations = allowed_locations
        else:
            unauthorized = set(selected_locations) - set(allowed_locations)
            if unauthorized:
                messages.warning(request, f"Access denied to: {', '.join(unauthorized)}")
                selected_locations = [loc for loc in selected_locations if loc in allowed_locations]
            
            if not selected_locations:
                selected_locations = allowed_locations
    
    if not selected_locations and not user_profile.is_admin:
        return HttpResponseForbidden("You don't have access to any locations.")
    
    # If admin selected 'all', reset to empty list
    if user_profile.is_admin and (not selected_locations or 'all' in request.GET.getlist('un_filter')):
        selected_locations = []
    
    selected_category = request.GET.get('category', 'all')
    selected_product = request.GET.get('prod_filter', 'all')
    selected_campaign = request.GET.get('campaign_filter', 'all')
    
    # Adjust dates to current year
    start_date = start_date.replace(year=current_year)
    end_date = end_date.replace(year=current_year)
    
    # Create timezone-aware datetimes
    start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
    end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
    
    def apply_filters(queryset):
        """Apply consistent filters across all queries"""
        if selected_locations:
            queryset = queryset.filter(un__in=selected_locations)
        if selected_category != 'all':
            queryset = queryset.filter(prodg=selected_category)
        if selected_product != 'all':
            queryset = queryset.filter(prod=selected_product)
        if selected_campaign != 'all':
            queryset = queryset.filter(actions=selected_campaign)
        return queryset.exclude(un__in=["მთავარი საწყობი 2", "სატესტო"])
    
    def get_year_stats(year):
        """Get comprehensive stats for a given year"""
        year_start = start_date.replace(year=year)
        year_end = end_date.replace(year=year)
        year_start_dt = timezone.make_aware(datetime.combine(year_start, datetime.min.time()))
        year_end_dt = timezone.make_aware(datetime.combine(year_end, datetime.max.time()))
        
        q = Sales.objects.filter(
            cd__year=year,
            cd__gte=year_start_dt,
            cd__lte=year_end_dt
        )
        q = apply_filters(q)
        
        # Basic stats
        basic_stats = q.aggregate(
            total_revenue=Sum('tanxa'),
            total_tickets=Count('zedd', distinct=True),
            total_items=Count('zedd'),
            discount_total=Sum('discount_price'),
            std_price_total=Sum('std_price')
        )
        
        # Calculate derived metrics
        total_tickets = basic_stats['total_tickets'] or 0
        total_revenue = float(basic_stats['total_revenue'] or 0)
        total_items = basic_stats['total_items'] or 0
        
        avg_basket = total_revenue / total_tickets if total_tickets > 0 else 0
        items_per_ticket = total_items / total_tickets if total_tickets > 0 else 0
        discount_share = (1 - (basic_stats['discount_total'] / basic_stats['std_price_total'])) * 100 if basic_stats['std_price_total'] and basic_stats['std_price_total'] > 0 else 0
        
        # Cross-selling stats
        ticket_items = q.filter(prodt='selling item').exclude(tanxa=0).exclude(prodg='POP').values('zedd').annotate(
            item_count=Count('idreal1')
        )
        
        total_analyzed_tickets = ticket_items.count()
        cross_sell_tickets = sum(1 for t in ticket_items if t['item_count'] >= 3)
        single_item_tickets = sum(1 for t in ticket_items if t['item_count'] == 1)
        
        cross_sell_rate = (cross_sell_tickets / total_analyzed_tickets * 100) if total_analyzed_tickets > 0 else 0
        single_item_rate = (single_item_tickets / total_analyzed_tickets * 100) if total_analyzed_tickets > 0 else 0
        
        # Category performance
        category_data = q.values('prodg').annotate(
            revenue=Sum('tanxa')
        ).order_by('-revenue')[:5]
        
        # Top products
        product_data = q.values('prod').annotate(
            revenue=Sum('tanxa'),
            quantity=Sum('raod')
        ).order_by('-revenue')[:5]
        
        # Location performance
        location_data = q.values('un').annotate(
            revenue=Sum('tanxa'),
            tickets=Count('zedd', distinct=True)
        ).order_by('-revenue')[:5]
        
        return {
            'year': year,
            'total_revenue': total_revenue,
            'total_tickets': total_tickets,
            'total_items': total_items,
            'avg_basket': avg_basket,
            'items_per_ticket': items_per_ticket,
            'discount_share': discount_share,
            'cross_sell_rate': cross_sell_rate,
            'single_item_rate': single_item_rate,
            'top_categories': list(category_data),
            'top_products': list(product_data),
            'top_locations': list(location_data)
        }
    
    # Get stats for all available years
    stats_current = get_year_stats(current_year)
    stats_previous = get_year_stats(previous_year)
    stats_two_years = get_year_stats(two_years_ago) if two_years_ago else None
    
    # Helper functions
    def calc_change(current, previous):
        if previous and previous > 0:
            return ((current - previous) / previous) * 100
        return 0
    
    def format_currency(value):
        if value >= 1000000:
            return f"${value/1000000:.1f}M"
        elif value >= 1000:
            return f"${value/1000:.1f}K"
        return f"${value:.2f}"
    
    def format_number(value):
        if value >= 1000000:
            return f"{value/1000000:.1f}M"
        elif value >= 1000:
            return f"{value/1000:.1f}K"
        return f"{int(value)}"
    
    def get_trend_class(change_pct):
        if change_pct > 0:
            return 'positive'
        elif change_pct < 0:
            return 'negative'
        return 'neutral'
    
    def get_trend_icon(change_pct):
        if change_pct > 0:
            return 'up'
        elif change_pct < 0:
            return 'down'
        return 'right'
    
    # Generate insights
    insights_list = []
    recommendations = []
    
    # Calculate changes
    revenue_change = calc_change(stats_current['total_revenue'], stats_previous['total_revenue'])
    tickets_change = calc_change(stats_current['total_tickets'], stats_previous['total_tickets'])
    basket_change = calc_change(stats_current['avg_basket'], stats_previous['avg_basket'])
    cross_sell_change = calc_change(stats_current['cross_sell_rate'], stats_previous['cross_sell_rate'])
    single_item_change = calc_change(stats_current['single_item_rate'], stats_previous['single_item_rate'])
    
    # INSIGHT 1: Overall Revenue Performance
    if abs(revenue_change) > 1:  # Only show if meaningful change
        revenue_insight = {
            'category': 'Revenue Analysis',
            'title': f"Revenue {'Growth' if revenue_change > 0 else 'Decline'} of {abs(revenue_change):.1f}%",
            'icon': 'fa-chart-line',
            'icon_class': 'icon-positive' if revenue_change > 0 else 'icon-negative',
            'description': '',
            'metrics': [
                {
                    'label': f'{current_year} Revenue',
                    'value': format_currency(stats_current['total_revenue']),
                    'change': f"{revenue_change:+.1f}%",
                    'change_class': get_trend_class(revenue_change),
                    'change_icon': get_trend_icon(revenue_change)
                },
                {
                    'label': f'{previous_year} Revenue',
                    'value': format_currency(stats_previous['total_revenue']),
                    'change': None
                }
            ],
            'year_comparison': None
        }
        
        # Generate description based on revenue components
        if revenue_change > 0:
            if tickets_change > basket_change:
                revenue_insight['description'] = f"<p>Your revenue increased by <span class='highlight-positive'>{revenue_change:.1f}%</span> compared to {previous_year}, primarily driven by a <strong>{tickets_change:.1f}% increase in transaction volume</strong>. This indicates strong customer acquisition or increased purchase frequency.</p>"
            else:
                revenue_insight['description'] = f"<p>Your revenue grew by <span class='highlight-positive'>{revenue_change:.1f}%</span> year-over-year, with the average basket size increasing by <strong>{basket_change:.1f}%</strong>. Customers are spending more per transaction, suggesting effective upselling or premium product adoption.</p>"
        else:
            revenue_insight['description'] = f"<p>Revenue declined by <span class='highlight-negative'>{abs(revenue_change):.1f}%</span> compared to {previous_year}. "
            if tickets_change < 0 and basket_change < 0:
                revenue_insight['description'] += f"Both transaction volume (down {abs(tickets_change):.1f}%) and average basket size (down {abs(basket_change):.1f}%) decreased, indicating challenges in both customer retention and purchase value.</p>"
            elif tickets_change < 0:
                revenue_insight['description'] += f"This is primarily due to a <strong>{abs(tickets_change):.1f}% decrease in transaction volume</strong>, despite average basket size remaining stable.</p>"
            else:
                revenue_insight['description'] += f"While transaction volume increased by {tickets_change:.1f}%, the average basket size decreased by {abs(basket_change):.1f}%, suggesting customers are purchasing less per visit.</p>"
        
        insights_list.append(revenue_insight)
        
        # Add recommendations based on revenue performance
        if revenue_change < 0:
            if tickets_change < -5:
                recommendations.append("Focus on customer retention and acquisition strategies to reverse the declining transaction volume. Consider loyalty programs or targeted marketing campaigns.")
            if basket_change < -5:
                recommendations.append("Implement bundle offers or cross-selling strategies to increase average basket size and maximize value per customer visit.")
    
    # INSIGHT 2: Cross-Selling Performance
    if stats_current['cross_sell_rate'] > 0:
        cross_sell_insight = {
            'category': 'Customer Behavior',
            'title': f"Cross-Selling Rate: {stats_current['cross_sell_rate']:.1f}%",
            'icon': 'fa-layer-group',
            'icon_class': 'icon-positive' if cross_sell_change > 0 else 'icon-warning',
            'description': '',
            'metrics': [
                {
                    'label': 'Cross-Sell Rate',
                    'value': f"{stats_current['cross_sell_rate']:.1f}%",
                    'change': f"{cross_sell_change:+.1f}%" if cross_sell_change != 0 else "No change",
                    'change_class': get_trend_class(cross_sell_change),
                    'change_icon': get_trend_icon(cross_sell_change)
                },
                {
                    'label': 'Single Item Rate',
                    'value': f"{stats_current['single_item_rate']:.1f}%",
                    'change': f"{single_item_change:+.1f}%" if single_item_change != 0 else "No change",
                    'change_class': 'negative' if single_item_change > 0 else 'positive',
                    'change_icon': get_trend_icon(single_item_change)
                }
            ],
            'year_comparison': None
        }
        
        if cross_sell_change > 5:
            cross_sell_insight['description'] = f"<p><strong>Excellent progress!</strong> Your cross-selling rate improved by <span class='highlight-positive'>{cross_sell_change:.1f}%</span>, with <strong>{stats_current['cross_sell_rate']:.1f}% of transactions</strong> containing 3+ items. This indicates effective merchandising and sales techniques.</p>"
            recommendations.append(f"Continue strengthening cross-selling initiatives. Consider training staff on successful bundling techniques and optimizing product placement.")
        elif cross_sell_change < -5:
            cross_sell_insight['description'] = f"<p>Cross-selling performance declined by <span class='highlight-negative'>{abs(cross_sell_change):.1f}%</span>. Only <strong>{stats_current['cross_sell_rate']:.1f}% of customers</strong> are purchasing 3+ items per transaction, down from {stats_previous['cross_sell_rate']:.1f}% last year.</p>"
            recommendations.append("Develop strategic product bundles and train staff on cross-selling techniques. Consider implementing 'frequently bought together' displays.")
        else:
            cross_sell_insight['description'] = f"<p>Your cross-selling rate is stable at <strong>{stats_current['cross_sell_rate']:.1f}%</strong>, with {format_number(stats_current['total_tickets'] * stats_current['cross_sell_rate'] / 100)} multi-item transactions. There's opportunity to further improve customer basket composition.</p>"
            
            if stats_current['single_item_rate'] > 30:
                cross_sell_insight['description'] += f"<p>However, <span class='highlight-warning'>{stats_current['single_item_rate']:.1f}% of transactions</span> are single-item purchases, representing a significant opportunity for improvement.</p>"
                recommendations.append(f"With {stats_current['single_item_rate']:.1f}% single-item purchases, focus on bundling strategies and point-of-sale suggestions to increase items per basket.")
        
        insights_list.append(cross_sell_insight)
    
    # INSIGHT 3: Basket Size Trends
    if abs(basket_change) > 3:
        basket_insight = {
            'category': 'Transaction Value',
            'title': f"Average Basket {'Increased' if basket_change > 0 else 'Decreased'} to ${stats_current['avg_basket']:.2f}",
            'icon': 'fa-shopping-basket',
            'icon_class': 'icon-positive' if basket_change > 0 else 'icon-negative',
            'description': '',
            'metrics': [
                {
                    'label': f'{current_year} Avg Basket',
                    'value': f"${stats_current['avg_basket']:.2f}",
                    'change': f"{basket_change:+.1f}%",
                    'change_class': get_trend_class(basket_change),
                    'change_icon': get_trend_icon(basket_change)
                },
                {
                    'label': 'Items per Ticket',
                    'value': f"{stats_current['items_per_ticket']:.1f}",
                    'change': None
                }
            ],
            'year_comparison': None
        }
        
        items_change = calc_change(stats_current['items_per_ticket'], stats_previous['items_per_ticket'])
        
        if basket_change > 0:
            if items_change > basket_change:
                basket_insight['description'] = f"<p>The average basket value increased by <span class='highlight-positive'>{basket_change:.1f}%</span> to <strong>${stats_current['avg_basket']:.2f}</strong>, primarily driven by customers purchasing more items per transaction (up {items_change:.1f}%).</p>"
            else:
                basket_insight['description'] = f"<p>Average basket size grew by <span class='highlight-positive'>{basket_change:.1f}%</span> to <strong>${stats_current['avg_basket']:.2f}</strong>, indicating customers are trading up to higher-value products or responding well to premium offerings.</p>"
                recommendations.append("Capitalize on the premium trend by highlighting high-margin products and creating exclusive bundles.")
        else:
            basket_insight['description'] = f"<p>The average basket decreased by <span class='highlight-negative'>{abs(basket_change):.1f}%</span> to <strong>${stats_current['avg_basket']:.2f}</strong>. "
            if items_change < 0:
                basket_insight['description'] += "Customers are purchasing fewer items per visit, suggesting potential issues with product availability, pricing, or shopping experience.</p>"
                recommendations.append("Investigate causes of smaller baskets - consider customer feedback surveys and analyze product availability during peak periods.")
            else:
                basket_insight['description'] += "While customers are buying similar quantities, they're choosing lower-priced options, possibly due to economic factors or competitive pricing pressure.</p>"
        
        insights_list.append(basket_insight)
    
    # INSIGHT 4: Category Performance (if available)
    if stats_current['top_categories']:
        top_cat = stats_current['top_categories'][0]
        top_cat_revenue = float(top_cat['revenue'] or 0)
        top_cat_share = (top_cat_revenue / stats_current['total_revenue'] * 100) if stats_current['total_revenue'] > 0 else 0
        
        # Find previous year data for same category
        prev_cat_data = next((c for c in stats_previous['top_categories'] if c['prodg'] == top_cat['prodg']), None)
        
        if prev_cat_data:
            prev_cat_revenue = float(prev_cat_data['revenue'] or 0)
            cat_change = calc_change(top_cat_revenue, prev_cat_revenue)
            
            category_insight = {
                'category': 'Category Performance',
                'title': f"{top_cat['prodg']} Leads with {top_cat_share:.1f}% Share",
                'icon': 'fa-tags',
                'icon_class': 'icon-positive' if cat_change > 0 else 'icon-warning',
                'description': f"<p><strong>{top_cat['prodg']}</strong> is your top-performing category, generating <span class='highlight'>{format_currency(top_cat_revenue)}</span> ({top_cat_share:.1f}% of total revenue). ",
                'metrics': [
                    {
                        'label': 'Category Revenue',
                        'value': format_currency(top_cat_revenue),
                        'change': f"{cat_change:+.1f}%",
                        'change_class': get_trend_class(cat_change),
                        'change_icon': get_trend_icon(cat_change)
                    },
                    {
                        'label': 'Revenue Share',
                        'value': f"{top_cat_share:.1f}%",
                        'change': None
                    }
                ],
                'year_comparison': None
            }
            
            if cat_change > 10:
                category_insight['description'] += f"This category grew by <span class='highlight-positive'>{cat_change:.1f}%</span> year-over-year, significantly outpacing overall business growth.</p>"
                recommendations.append(f"Invest in expanding the {top_cat['prodg']} category - increase inventory depth, add complementary products, and feature prominently in marketing.")
            elif cat_change < -10:
                category_insight['description'] += f"However, this category declined by <span class='highlight-negative'>{abs(cat_change):.1f}%</span> compared to last year, which is concerning given its importance to your business.</p>"
                recommendations.append(f"Investigate the decline in {top_cat['prodg']} - analyze pricing, competition, and product freshness. Consider category refresh or promotional support.")
            else:
                category_insight['description'] += f"Performance changed by {cat_change:+.1f}% versus last year.</p>"
            
            insights_list.append(category_insight)
    
    # INSIGHT 5: Location Performance (if filtered or if there's variance)
    if stats_current['top_locations'] and len(stats_current['top_locations']) > 1:
        top_loc = stats_current['top_locations'][0]
        bottom_loc = stats_current['top_locations'][-1]
        
        top_loc_revenue = float(top_loc['revenue'] or 0)
        bottom_loc_revenue = float(bottom_loc['revenue'] or 0)
        
        if top_loc_revenue > 0 and bottom_loc_revenue > 0:
            variance_ratio = top_loc_revenue / bottom_loc_revenue
            
            if variance_ratio > 2:  # Significant variance
                location_insight = {
                    'category': 'Location Analysis',
                    'title': 'Significant Performance Variance Across Locations',
                    'icon': 'fa-map-marker-alt',
                    'icon_class': 'icon-warning',
                    'description': f"<p>There's significant performance variance across locations. <strong>{top_loc['un']}</strong> generates {format_currency(top_loc_revenue)}, while <strong>{bottom_loc['un']}</strong> generates {format_currency(bottom_loc_revenue)} - a {variance_ratio:.1f}x difference.</p>",
                    'metrics': [
                        {
                            'label': 'Top Location',
                            'value': format_currency(top_loc_revenue),
                            'change': None
                        },
                        {
                            'label': 'Performance Spread',
                            'value': f"{variance_ratio:.1f}x",
                            'change': None
                        }
                    ],
                    'year_comparison': None
                }
                
                recommendations.append(f"Analyze best practices from {top_loc['un']} and apply learnings to underperforming locations. Consider staffing, inventory, and local marketing differences.")
                insights_list.append(location_insight)
    
    # INSIGHT 6: Multi-year trend (if we have 3 years of data)
    if stats_two_years:
        revenue_3yr_growth = calc_change(stats_current['total_revenue'], stats_two_years['total_revenue'])
        cagr = (((stats_current['total_revenue'] / stats_two_years['total_revenue']) ** (1/2)) - 1) * 100 if stats_two_years['total_revenue'] > 0 else 0
        
        if abs(revenue_3yr_growth) > 10:
            trend_insight = {
                'category': 'Long-term Trends',
                'title': f"{current_year - two_years_ago}-Year Performance Trajectory",
                'icon': 'fa-chart-area',
                'icon_class': 'icon-positive' if revenue_3yr_growth > 0 else 'icon-negative',
                'description': f"<p>Over the past {current_year - two_years_ago} years, revenue {'grew' if revenue_3yr_growth > 0 else 'declined'} by <span class='{'highlight-positive' if revenue_3yr_growth > 0 else 'highlight-negative'}'>{abs(revenue_3yr_growth):.1f}%</span> (CAGR: {cagr:+.1f}%). ",
                'metrics': [],
                'year_comparison': [
                    {
                        'year': str(current_year),
                        'stats': [
                            {'label': 'Revenue', 'value': format_currency(stats_current['total_revenue'])},
                            {'label': 'Tickets', 'value': format_number(stats_current['total_tickets'])},
                            {'label': 'Avg Basket', 'value': f"${stats_current['avg_basket']:.2f}"}
                        ]
                    },
                    {
                        'year': str(previous_year),
                        'stats': [
                            {'label': 'Revenue', 'value': format_currency(stats_previous['total_revenue'])},
                            {'label': 'Tickets', 'value': format_number(stats_previous['total_tickets'])},
                            {'label': 'Avg Basket', 'value': f"${stats_previous['avg_basket']:.2f}"}
                        ]
                    },
                    {
                        'year': str(two_years_ago),
                        'stats': [
                            {'label': 'Revenue', 'value': format_currency(stats_two_years['total_revenue'])},
                            {'label': 'Tickets', 'value': format_number(stats_two_years['total_tickets'])},
                            {'label': 'Avg Basket', 'value': f"${stats_two_years['avg_basket']:.2f}"}
                        ]
                    }
                ]
            }
            
            # Analyze the trend trajectory
            recent_growth = calc_change(stats_current['total_revenue'], stats_previous['total_revenue'])
            older_growth = calc_change(stats_previous['total_revenue'], stats_two_years['total_revenue'])
            
            if recent_growth > older_growth:
                trend_insight['description'] += f"Growth is <strong>accelerating</strong> - {current_year} saw {recent_growth:.1f}% growth compared to {older_growth:.1f}% in the prior year.</p>"
            elif recent_growth < older_growth:
                trend_insight['description'] += f"Growth is <strong>decelerating</strong> - {current_year} saw {recent_growth:.1f}% growth compared to {older_growth:.1f}% in the prior year.</p>"
            else:
                trend_insight['description'] += f"Growth is <strong>consistent</strong> at approximately {recent_growth:.1f}% year-over-year.</p>"
            
            insights_list.append(trend_insight)
    
    # Prepare summary for the overview section
    summary = {
        'total_revenue': format_currency(stats_current['total_revenue']),
        'revenue_change': f"{abs(revenue_change):.1f}%",
        'revenue_trend': get_trend_class(revenue_change),
        'revenue_trend_icon': get_trend_icon(revenue_change),
        
        'total_tickets': format_number(stats_current['total_tickets']),
        'tickets_change': f"{abs(tickets_change):.1f}%",
        'tickets_trend': get_trend_class(tickets_change),
        'tickets_trend_icon': get_trend_icon(tickets_change),
        
        'avg_basket': f"${stats_current['avg_basket']:.2f}",
        'basket_change': f"{abs(basket_change):.1f}%",
        'basket_trend': get_trend_class(basket_change),
        'basket_trend_icon': get_trend_icon(basket_change),
        
        'cross_sell_rate': f"{stats_current['cross_sell_rate']:.1f}",
        'cross_sell_change': f"{abs(cross_sell_change):.1f}%",
        'cross_sell_trend': get_trend_class(cross_sell_change),
        'cross_sell_trend_icon': get_trend_icon(cross_sell_change)
    }
    
    # Date range text
    date_range_text = f"{start_date.strftime('%b %d')} - {end_date.strftime('%b %d')}, {current_year}"
    if selected_locations:
        location_text = ', '.join(selected_locations[:3])
        if len(selected_locations) > 3:
            location_text += f" +{len(selected_locations) - 3} more"
        date_range_text += f" • {location_text}"
    
    context = {
        'insights': insights_list,
        'recommendations': recommendations,
        'summary': summary,
        'date_range_text': date_range_text,
        'current_year': current_year,
        'previous_year': previous_year,
        'two_years_ago': two_years_ago,

        'user_profile': user_profile,
        'is_admin': user_profile.is_admin
    }
    
    return render(request, 'insights.html', context)

def health(request):
    return HttpResponse("ok")

@login_required
def stat_main(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            WITH base AS (
                SELECT "UN", SUM("Tanxa") AS total
                FROM sales_main_web
                WHERE extract(year from "CD") = 2026
                GROUP BY "UN"
            )
            SELECT * FROM base ORDER BY total DESC LIMIT 20;
        """)
        rows = cursor.fetchall()
    
    context = {
        'location': rows
    }
    return render(request, 'stat_main.html', context)

@login_required
def get_filter_options(request):
    """AJAX endpoint to get available filter options based on current selections"""
    try:
        user_profile = request.user.profile
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    allowed_locations = user_profile.get_allowed_locations()
    current_year = int(request.GET.get('year', 2026))
    selected_locations = request.GET.getlist('un_filter')
    selected_category = request.GET.get('category', 'all')
    selected_product = request.GET.get('prod_filter', 'all')
    selected_campaign = request.GET.get('campaign_filter', 'all')
    
    base_query = Sales.objects.filter(cd__year=current_year).exclude(
        un__in=["მთავარი საწყობი 2", "სატესტო"]
    )
    
    # LOCATIONS
    q_locations = base_query
    if selected_category != 'all':
        q_locations = q_locations.filter(prodg=selected_category)
    if selected_product != 'all':
        q_locations = q_locations.filter(prod=selected_product)
    if selected_campaign != 'all':
        q_locations = q_locations.filter(actions=selected_campaign)
    available_locations = list(q_locations.values_list('un', flat=True).distinct().order_by('un'))
    
    # CATEGORIES
    q_categories = base_query
    if selected_locations:
        q_categories = q_categories.filter(un__in=selected_locations)
    if selected_product != 'all':
        q_categories = q_categories.filter(prod=selected_product)
    if selected_campaign != 'all':
        q_categories = q_categories.filter(actions=selected_campaign)
    available_categories = list(q_categories.values_list('prodg', flat=True).distinct().order_by('prodg'))
    
    # PRODUCTS
    q_products = base_query
    if selected_locations:
        q_products = q_products.filter(un__in=selected_locations)
    if selected_category != 'all':
        q_products = q_products.filter(prodg=selected_category)
    if selected_campaign != 'all':
        q_products = q_products.filter(actions=selected_campaign)
    available_products = list(q_products.values_list('prod', flat=True).distinct().order_by('prod'))
    
    # CAMPAIGNS
    q_campaigns = base_query
    if selected_locations:
        q_campaigns = q_campaigns.filter(un__in=selected_locations)
    if selected_category != 'all':
        q_campaigns = q_campaigns.filter(prodg=selected_category)
    if selected_product != 'all':
        q_campaigns = q_campaigns.filter(prod=selected_product)
    available_campaigns = list(q_campaigns.values_list('actions', flat=True).distinct().order_by('actions'))
    
    return JsonResponse({
        'locations': available_locations if user_profile.is_admin else allowed_locations,
        'categories': available_categories,
        'products': available_products,
        'campaigns': available_campaigns
    })

def competitive(request):
    try:
        user_profile = request.user.profile
    except:
        return HttpResponseForbidden("Access denied. Contact administrator.")
    
    # ADMIN ONLY for query interface
    if not user_profile.is_admin:
        return HttpResponseForbidden("Only administrators can access the SQL query interface.")
    GLOW_CODES = [
        "62903"
            ]
        # Plan quantities per location (AAG target)
    PLAN_QTY = {
        "ბათუმი გრანდ მოლი": 12,
        "ბათუმი მეტრო მოლი": 6,
        "გალერია":18,
        "გლდანი": 16,
        "გლდანი სითი მოლი": 8,
        "გუდვილი": 10,
        "გუდვილი 2 ": 20,
        "ვაკე 1": 8,
        "ისტ პოინტი": 16,
        "მერანი": 4,
        "პეკინი ": 14,
        "პლეხანოვი ": 10,
        "რუსთავი": 8,
    }
    
    start_date = request.GET.get('start_date', '2026-05-01')
    end_date = request.GET.get('end_date', '2026-05-31')
    
    with connection.cursor() as cursor:
        
        # Skincare by location
        cursor.execute("""
            SELECT 
                "UN" as location,
                SUM(CASE WHEN "ProdG" = 'PARFUMS' THEN "Tanxa" ELSE 0 END) as skincare_revenue,
                SUM("Tanxa") as total_revenue,
                ROUND(
                    (100.0 * SUM(CASE WHEN "ProdG" = 'PARFUMS' THEN "Tanxa" ELSE 0 END) 
                    / NULLIF(SUM("Tanxa"), 0))::numeric, 2
                ) as skincare_pct
            FROM sales_main_web
            WHERE "CD" >= %s AND "CD" <= %s and "UN" <> 'გორი' and "ProdG" <> 'POP'
            AND "Tanxa" != 0
            GROUP BY "UN"
            ORDER BY skincare_pct DESC
        """, [start_date, end_date])
        skincare_rows = cursor.fetchall()

        # Glow by location — revenue, quantity, share
        cursor.execute("""
            SELECT 
                "UN" as location,
                SUM(CASE WHEN left(right("IdProd",6),5) IN %s THEN "Tanxa" ELSE 0 END) as glow_revenue,
                SUM(CASE WHEN left(right("IdProd",6),5) IN %s THEN 1 ELSE 0 END) as glow_qty,
                SUM("Tanxa") as total_revenue,
                ROUND(
                    (100.0 * SUM(CASE WHEN left(right("IdProd",6),5) IN %s THEN "Tanxa" ELSE 0 END) 
                    / NULLIF(SUM("Tanxa"), 0))::numeric, 2
                ) as glow_pct
            FROM sales_main_web
            WHERE "CD" >= %s AND "CD" <= %s and "UN" <> 'გორი' and "ProdG" <> 'POP'
            AND "Tanxa" != 0
            GROUP BY "UN"
            ORDER BY glow_revenue DESC
        """, [tuple(GLOW_CODES), tuple(GLOW_CODES), tuple(GLOW_CODES), start_date, end_date])
        glow_rows = cursor.fetchall()
        # Add plan + achievement %
        glow_rows_enriched = []

        for row in glow_rows:
            location = row[0]
            actual_qty = row[2]

            plan = PLAN_QTY.get(location, 0)

            achievement = round((actual_qty / plan) * 100, 1) if plan > 0 else 0

            glow_rows_enriched.append((
                *row,          # existing data
                plan,          # index 5
                achievement    # index 6
            ))

        # Glow per product breakdown
        cursor.execute("""
            SELECT 
                "Prod" as product_name,
                "IdProd" as product_code,
                SUM("Tanxa") as revenue,
                SUM("raod") as quantity,
                COUNT(DISTINCT "Zedd") as tickets
            FROM sales_main_web
            WHERE "CD" >= %s AND "CD" <= %s and "UN" <> 'გორი'
            AND left(right("IdProd",6),5) IN %s
            AND "Tanxa" != 0
            GROUP BY "Prod", "IdProd"
            ORDER BY revenue DESC
        """, [start_date, end_date, tuple(GLOW_CODES)])
        glow_products = cursor.fetchall()

    # Best glow location
    best_glow_location = glow_rows[0] if glow_rows else None
    # After glow_rows_enriched is built:
    total_glow_revenue = sum(r[1] for r in glow_rows_enriched)
    total_glow_qty = sum(r[2] for r in glow_rows_enriched)
    total_glow_plan = sum(r[5] for r in glow_rows_enriched)
    total_glow_achievement = round((total_glow_qty / total_glow_plan) * 100, 1) if total_glow_plan > 0 else 0

    # After skincare_rows:
    total_skincare_revenue = sum(r[1] for r in skincare_rows)
    total_all_revenue = sum(r[2] for r in skincare_rows)
    total_skincare_pct = round((total_skincare_revenue / total_all_revenue) * 100, 2) if total_all_revenue > 0 else 0

    # After glow_products:
    total_prod_revenue = sum(p[2] for p in glow_products)
    total_prod_qty = sum(p[3] for p in glow_products)
    total_prod_tickets = sum(p[4] for p in glow_products)
    context = {
        'skincare_rows': skincare_rows,
        'glow_rows': glow_rows_enriched,
        'glow_products': glow_products,
        'best_glow_location': best_glow_location,
        'start_date': start_date,
        'end_date': end_date,
        'total_skincare_revenue': total_skincare_revenue,
        'total_all_revenue': total_all_revenue,
        'total_skincare_pct': total_skincare_pct,
        'total_glow_revenue': total_glow_revenue,
        'total_glow_qty': total_glow_qty,
        'total_glow_plan': total_glow_plan,
        'total_glow_achievement': total_glow_achievement,
        'total_prod_revenue': total_prod_revenue,
        'total_prod_qty': total_prod_qty,
        'total_prod_tickets': total_prod_tickets,
    }
    return render(request, 'competition_motivation.html', context)

def bonus(request):
    try:
        user_profile = request.user.profile
    except:
        return HttpResponseForbidden("Access denied. Contact administrator.")

    if not user_profile.is_admin:
        return HttpResponseForbidden("Only administrators can access the SQL query interface.")

    year = request.GET.get('year', 2026)
    month = request.GET.get('month', 6)
    working_days = request.GET.get('working_days', 30)

    with connection.cursor() as cursor:
        cursor.execute("""
                      with date_filtered_base_cte as (
    select 
        "Tanam", 
        "ProdG", 
        "Tanxa", 
        "Zedd",
        "UN"
    from sales_main_web
    where extract(year from "CD") = %s
      and extract(month from "CD") = %s
),


plan_cte as (
    select *
    from (values 
    ('გუდვილი 2 ', 113000),
    ('გალერია', 82000),
    ('გლდანი', 58000),
    ('პეკინი ', 55000),
    ('ისტ პოინტი', 57000),
    ('გლდანი სითი მოლი', 33000),
    ('პლეხანოვი ', 43000),
    ('ბათუმი მეტრო მოლი', 45000),
    ('რუსთავი', 31000),
    ('გუდვილი', 46000),
    ('ვაკე 1', 43000),
    ('ბათუმი გრანდ მოლი', 58000),
    ('მერანი', 16000)
    ) as t("UN", plan)
),

vacation_cte as (
    select * from (values
        ('მარი ლომიძე', 1),
        ('მარიამ მანტაშიან', 2),
        ('მარიამ გვარამაძე', 1),
        ('ნათია ჯანგულაშვილი', 2),
        ('თეონა თევდორაშვილი', 3),
        ('თაკო დათუაშვილი', 5),
        ('მაგდა მაგდალიანოვა', 2),
        ('სოფო ფილია', 1),
        ('ლიკა დარისპანაშვილი', 2),
        ('ეკატერინე კუტალაძე', 4),
        ('მარიამ კონიაშვილი', 1),
        ('ვერიკო გვასალია', 3),
        ('ეკა ჟუჟუნაშილი', 4),
        ('მარიამ როსტომაშვილი', 2),
        ('მეგი თურმანიძე', 2),
        ('თამუნა ტაბატაძე', 2),
        ('ნინო ბერიძე', 2),
        ('სოფიო ფევაძე', 8),
        ('ია აბულაძე', 2),
        ('ქეთევან ბესელია', 3),
        ('მარი ჟვანია', 5),
        ('ეკატერინე ჩოხელი', 3),
        ('ქეთი რუხაძე', 3),
        ('ქრისტინე გულბიანი', 6),
        ('რუსო ინასარიძე', 10),
        ('ხათუნა კვიწინაძე 1', 2)
    ) as t("Tanam", days_absent)
),

working_days_cte as (
    select %s as working_days
),

managers_not_cte as (
    select *,
    case 
        when trim("Tanam") in (
            'ნინო ბერიძე',
            'შორენა კვარაცხელია',
            'მანანა თეთრაძე',
            'ხათუნა კვიწინაძე 1',
            'მარიამ კონიაშვილი',
            'ნინო ართილაყვა',
            'ნათია ჯანგულაშვილი',
            'მაგდა მაგდალიანოვა',
            'მარი ჟვანია',
            'ნატა გურეშიძე',
            'თამუნა გელაძე',
            'ქეთევან ინასარიძე',
            'სოფიკო ფარფალია'
        ) 
    then 1 else 0 end as manager
    from date_filtered_base_cte
    where trim("Tanam") not in (select "UN" from plan_cte)
    and trim("Tanam") <> trim("UN") and trim("Tanam") <> '-'
),

cte_zedd_by_un as (
    select "UN",
        sum("Tanxa") as zedd_total
    from sales_main_web
    where extract(year from "CD") = 2026
      and extract(month from "CD") = 6
      and length("Zedd") = 10
    group by "UN"
),

cte_all_employee_count as (
    select "UN",
        count(distinct "Tanam") as all_employee_count
    from managers_not_cte
    group by "UN"
),

cte_zedd_per_employee as (
    select
        z."UN",
        round((z.zedd_total / NULLIF(ec.all_employee_count, 0))::numeric, 3) as zedd_share
    from cte_zedd_by_un z
    left join cte_all_employee_count ec on ec."UN" = z."UN"
),

cte_skincare_per as (
    select "Tanam", "manager", "UN",
        case 
            when "manager" = 0 
            then sum(case when "ProdG" = 'SKIN CARE' then "Tanxa" else 0 end) 
                / NULLIF(sum(case when "ProdG" <> 'POP' then "Tanxa" else 0 end), 0)
            else 0 
        end as skincare_percentage,
        sum("Tanxa") as total_turnover
    from managers_not_cte
    group by "Tanam", "manager", "UN"
),

cte_un_turnover as (
    select "UN",
        sum("Tanxa") as un_total_turnover
    from date_filtered_base_cte
    group by "UN"
),

cte_cross_base as (
    select "Tanam", "Zedd", "manager", "UN",
        case when count("Zedd") >= 3 then 1 else 0 end as counter_total_3_or_more
    from managers_not_cte
    where "ProdG" <> 'POP'
    group by "Tanam", "Zedd", "manager", "UN"
),

cte_un_cross as (
    select 
        "UN",
        sum(counter_total_3_or_more * 1.0) / NULLIF(count("Zedd"), 0) as un_cross_selling_percentage
    from cte_cross_base
    group by "UN"
),

cte_cross_main as (
    select 
        cb."Tanam",
        case 
            when cb.manager = 0 
            then sum(cb.counter_total_3_or_more * 1.0) / NULLIF(count(cb."Zedd"), 0)
            else max(uc.un_cross_selling_percentage)
        end as cross_selling_percentage
    from cte_cross_base cb
    left join cte_un_cross uc on uc."UN" = cb."UN"
    group by cb."Tanam", cb.manager
),

cte_consultant_count as (
    select "UN",
        count(distinct "Tanam") as consultant_count
    from managers_not_cte
    where manager = 0
    group by "UN"
),

cte_total_headcount as (
    select "UN",
        count(distinct "Tanam") as total_headcount
    from managers_not_cte
    group by "UN"
),

total_base_per_employee as (
    select  
        c."Tanam",
        c."manager",
        c."UN",
        round(c."skincare_percentage"::numeric, 3) as skincare_percentage,
        round(case
            when c."manager" = 1 then ut.un_total_turnover
            else c.total_turnover                                          -- own sales only
        end::numeric, 3) as total_turnover,
        round(case
            when c."manager" = 1 then ut.un_total_turnover + coalesce(ze.zedd_share, 0)
            else c.total_turnover + coalesce(ze.zedd_share, 0)            -- own sales + zedd only
        end::numeric, 3) as total_turnover_with_zedd,
        round(case
            when c."manager" = 1 then c.total_turnover / NULLIF(hc.total_headcount, 0)
            else c.total_turnover                                          -- own sales only
        end::numeric, 3) as bonus_turnover,
        coalesce(ze.zedd_share, 0) as zedd_share,
        round(c1.cross_selling_percentage::numeric, 3) as cross_selling_percentage,
        p.plan,
        case 
            when c."manager" = 1 then p.plan
            else round((p.plan * 0.9 / NULLIF(cc.consultant_count, 0))::numeric, 3)
        end as individual_plan
    from cte_skincare_per c
    left join cte_cross_main c1 on c1."Tanam" = c."Tanam"
    left join plan_cte p on p."UN" = c."UN"
    left join cte_consultant_count cc on cc."UN" = c."UN"
    left join cte_zedd_per_employee ze on ze."UN" = c."UN"
    left join cte_total_headcount hc on hc."UN" = c."UN"
    left join cte_skincare_per mgr on mgr."UN" = c."UN" and mgr."manager" = 1
    left join cte_un_turnover ut on ut."UN" = c."UN"
),

calculated_steps_fixed as (
    select *,
        case
            when manager = 0 then 
                case
                    when skincare_percentage >= 0.2 and skincare_percentage < 0.25 then 50 
                    when skincare_percentage >= 0.25 then 100 
                    else 0 
                end
            else 0 
        end AS skincare_fixed_cost,
        case
            when manager = 0 and cross_selling_percentage >= 0.4 then 50 
            else 0 
        end AS cross_selling_fixed_cost_consultant,
        case
            when manager = 1 and cross_selling_percentage >= 0.35 then 100 
            else 0 
        end AS cross_selling_fixed_cost_manager,
        case
            when manager = 0 then 
                case
                    when (total_turnover_with_zedd / NULLIF(individual_plan, 0)) >= 1.1 then 150
                    when (total_turnover_with_zedd / NULLIF(individual_plan, 0)) >= 1 then 100
                    else 0 
                end
            else 0 
        end AS total_fixed_individual_achievement
    FROM total_base_per_employee
),

cte_un_achievement as (
    select
        ut."UN",
        ut.un_total_turnover,
        p.plan,
        (ut.un_total_turnover / NULLIF(p.plan::numeric, 0)) as un_achievement_ratio,
        case
            when ut.un_total_turnover / NULLIF(p.plan::numeric, 0) >= 1.1 then 2.1
            when ut.un_total_turnover / NULLIF(p.plan::numeric, 0) >= 0.95 then 2.0
            when ut.un_total_turnover / NULLIF(p.plan::numeric, 0) >= 0.85 then 1.7
            else 0
        end as un_achievement_multiplier
    from cte_un_turnover ut
    left join plan_cte p on p."UN" = ut."UN"
),

cte_main_bonus as (
    select
        csf.*,
        round(ua.un_achievement_ratio::numeric, 3) as un_achievement_ratio,
        ua.un_achievement_multiplier,
        case
            when csf.manager = 0 then
                round(
                    ((ua.un_total_turnover * 0.01 / NULLIF(cc.consultant_count + 1, 0))
                    * ua.un_achievement_multiplier)::numeric
                , 3)
            else
                round(
                    ((ua.un_total_turnover * 0.01 / NULLIF(cc.consultant_count + 1, 0)) * 1.1
                    * ua.un_achievement_multiplier)::numeric
                , 3)
        end as main_bonus
    from calculated_steps_fixed csf
    left join cte_un_achievement ua on ua."UN" = csf."UN"
    left join cte_consultant_count cc on cc."UN" = csf."UN"
),

cte_vacation_adjusted as (
    select
        mb.*,
        coalesce(v.days_absent, 0) as days_absent,
        wd.working_days,
        round(
            (mb.main_bonus * coalesce(v.days_absent, 0)::numeric / wd.working_days)
        , 3) as vacation_deduction
    from cte_main_bonus mb
    cross join working_days_cte wd
    left join vacation_cte v on trim(v."Tanam") = trim(mb."Tanam")
),

cte_un_employee_count as (
    select "UN",
        count(*) as total_employees
    from cte_vacation_adjusted
    group by "UN"
),

cte_redistribution as (
    select
        recipient."UN",
        recipient."Tanam",
        round(sum(
            donor.vacation_deduction / NULLIF(ec.total_employees - 1, 0)
        )::numeric, 3) as redistribution_received
    from cte_vacation_adjusted recipient
    join cte_vacation_adjusted donor
        on donor."UN" = recipient."UN"
        and donor."Tanam" <> recipient."Tanam"
        and donor.vacation_deduction > 0
    join cte_un_employee_count ec on ec."UN" = donor."UN"
    group by recipient."UN", recipient."Tanam"
)

SELECT
    va.*,
    coalesce(r.redistribution_received, 0) as redistribution_received,
    round((
        va.main_bonus
        - va.vacation_deduction
        + coalesce(r.redistribution_received, 0)
    )::numeric, 3) as final_main_bonus,
    round((
        va.skincare_fixed_cost +
        va.cross_selling_fixed_cost_consultant +
        va.cross_selling_fixed_cost_manager +
        va.total_fixed_individual_achievement +
        va.main_bonus
        - va.vacation_deduction
        + coalesce(r.redistribution_received, 0)
    )::numeric, 3) AS total_bonus
FROM cte_vacation_adjusted va
left join cte_redistribution r
    on r."UN" = va."UN"
    and r."Tanam" = va."Tanam"
ORDER BY va."UN"
        """, [year, month, working_days])

        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # ── Build subtotals per UN ──────────────────────────────────────────
    from collections import defaultdict
    subtotals = defaultdict(lambda: {
        'skincare': 0.0, 'cross': 0.0, 'indiv': 0.0,
        'main_bonus': 0.0, 'vac_deduction': 0.0,
        'redistribution': 0.0, 'total_bonus': 0.0, 'count': 0,
    })
    grand = {
        'skincare': 0.0, 'cross': 0.0, 'indiv': 0.0,
        'main_bonus': 0.0, 'vac_deduction': 0.0,
        'redistribution': 0.0, 'total_bonus': 0.0, 'count': 0,
    }
    for r in rows:
        un = r['UN']
        sc  = float(r.get('skincare_fixed_cost') or 0)
        cr  = float(r.get('cross_selling_fixed_cost_consultant') or 0) + float(r.get('cross_selling_fixed_cost_manager') or 0)
        ind = float(r.get('total_fixed_individual_achievement') or 0)
        mb  = float(r.get('main_bonus') or 0)
        vd  = float(r.get('vacation_deduction') or 0)
        rd  = float(r.get('redistribution_received') or 0)
        tb  = float(r.get('total_bonus') or 0)
        subtotals[un]['skincare']       += sc
        subtotals[un]['cross']          += cr
        subtotals[un]['indiv']          += ind
        subtotals[un]['main_bonus']     += mb
        subtotals[un]['vac_deduction']  += vd
        subtotals[un]['redistribution'] += rd
        subtotals[un]['total_bonus']    += tb
        subtotals[un]['count']          += 1
        grand['skincare']       += sc
        grand['cross']          += cr
        grand['indiv']          += ind
        grand['main_bonus']     += mb
        grand['vac_deduction']  += vd
        grand['redistribution'] += rd
        grand['total_bonus']    += tb
        grand['count']          += 1

    return render(request, 'bonus.html', {
        'location': rows,
        'subtotals': dict(subtotals),
        'grand': grand,
        'working_days': working_days,
    })
# mmm

def _log_log_regression(prices, quantities):
    """
    Simple OLS log-log regression: ln(Q) = a + b*ln(P)
    Returns (elasticity, r_squared, intercept)
    Elasticity b < -1  → elastic
    Elasticity b in (-1, 0) → inelastic
    """
    pairs = [(p, q) for p, q in zip(prices, quantities) if p > 0 and q > 0]
    n = len(pairs)
    if n < 2:
        return None, None, None

    lp = [math.log(p) for p, _ in pairs]
    lq = [math.log(q) for _, q in pairs]

    mean_lp = sum(lp) / n
    mean_lq = sum(lq) / n

    cov = sum((lp[i] - mean_lp) * (lq[i] - mean_lq) for i in range(n))
    var = sum((lp[i] - mean_lp) ** 2 for i in range(n))

    if var == 0:
        return None, None, None

    b = cov / var
    a = mean_lq - b * mean_lp

    # R²
    ss_res = sum((lq[i] - (a + b * lp[i])) ** 2 for i in range(n))
    ss_tot = sum((lq[i] - mean_lq) ** 2 for i in range(n))
    r2 = 1 - ss_res / ss_tot if ss_tot != 0 else 0

    return round(b, 4), round(r2, 4), round(a, 4)

@staff_member_required(login_url='/login/')
def admin_main_only_me(request):
    # ── Filters from GET ──────────────────────────────────────────────
    selected_group  = request.GET.get('prodg', '')
    selected_product = request.GET.get('prod', '')
    price_type      = request.GET.get('price_type', 'std_price')   # std_price | discount_price
    demand_metric   = request.GET.get('demand_metric', 'quantity')  # quantity | tickets
    selected_location = request.GET.get('location', 'all')

    # ── Dropdown options ──────────────────────────────────────────────
    all_groups = (
        Sales.objects.exclude(prodg__isnull=True)
        .exclude(prodg='')
        .values_list('prodg', flat=True)
        .distinct()
        .order_by('prodg')
    )

    # Products filtered by group if chosen
    products_qs = Sales.objects.exclude(prod__isnull=True).exclude(prod='')
    if selected_group:
        products_qs = products_qs.filter(prodg=selected_group)
    all_products = products_qs.values_list('prod', flat=True).distinct().order_by('prod')

    all_locations = (
        Sales.objects.exclude(un__isnull=True)
        .exclude(un='')
        .values_list('un', flat=True)
        .distinct()
        .order_by('un')
    )

    # ── Base queryset ─────────────────────────────────────────────────
    qs = Sales.objects.all()
    if selected_location != 'all':
        qs = qs.filter(un=selected_location)
    if selected_group:
        qs = qs.filter(prodg=selected_group)
    if selected_product:
        qs = qs.filter(prod=selected_product)

    # ── Aggregate by price point ──────────────────────────────────────
    price_field = 'std_price' if price_type == 'std_price' else 'discount_price'

    # We need: for each distinct price → total demand
    price_demand_qs = (
        qs
        .exclude(**{f'{price_field}__isnull': True})
        .exclude(**{f'{price_field}__lte': 0})
        .values(price_field)
        .annotate(
            total_quantity=Sum('raod'),
            total_tickets=Count('zedd'),
            total_revenue=Sum('tanxa'),
            avg_discount=Avg('discount_price'),
        )
        .order_by(price_field)
    )

    rows = list(price_demand_qs)

    # Pick demand column
    def get_demand(row):
        if demand_metric == 'quantity':
            v = row.get('total_quantity') or 0
        else:
            v = row.get('total_tickets') or 0
        return float(v)

    prices    = [float(r[price_field]) for r in rows]
    demands   = [get_demand(r) for r in rows]
    revenues  = [float(r['total_revenue'] or 0) for r in rows]

    # ── Elasticity calculation ────────────────────────────────────────
    elasticity, r_squared, intercept = _log_log_regression(prices, demands)

    # Classify
    if elasticity is not None:
        abs_e = abs(elasticity)
        if abs_e > 1.5:
            elasticity_label = "HIGHLY ELASTIC"
            elasticity_color = "#ef4444"
            elasticity_advice = "Very price-sensitive. Small price increases will sharply reduce demand. Consider volume-based strategy."
        elif abs_e > 1.0:
            elasticity_label = "ELASTIC"
            elasticity_color = "#fb923c"
            elasticity_advice = "Customers react noticeably to price changes. Promotions will drive meaningful volume lift."
        elif abs_e > 0.5:
            elasticity_label = "INELASTIC"
            elasticity_color = "#10b981"
            elasticity_advice = "Customers are relatively price-tolerant. You have moderate pricing power."
        else:
            elasticity_label = "HIGHLY INELASTIC"
            elasticity_color = "#06b6d4"
            elasticity_advice = "Strong pricing power. Customers are not sensitive to price — loyalty or necessity driver."
    else:
        elasticity_label = "INSUFFICIENT DATA"
        elasticity_color = "#94a3b8"
        elasticity_advice = "Not enough price variation in the data. Select a product with more price history."

    # Fitted demand curve for chart overlay
    fitted = []
    if elasticity is not None and intercept is not None:
        for p in prices:
            if p > 0:
                fitted.append(round(math.exp(intercept + elasticity * math.log(p)), 2))
            else:
                fitted.append(None)

    # ── Point elasticities between consecutive price points ───────────
    point_elasticities = []
    for i in range(1, len(prices)):
        p1, p2 = prices[i-1], prices[i]
        d1, d2 = demands[i-1], demands[i]
        if p1 > 0 and d1 > 0 and p2 != p1 and d2 != d1:
            pct_d = (d2 - d1) / d1 * 100
            pct_p = (p2 - p1) / p1 * 100
            pe = round(pct_d / pct_p, 3) if pct_p != 0 else None
        else:
            pe = None
        point_elasticities.append({
            'price_from': round(p1, 2),
            'price_to':   round(p2, 2),
            'demand_from': round(d1, 2),
            'demand_to':   round(d2, 2),
            'pct_price': round((p2 - p1) / p1 * 100, 2) if p1 > 0 else None,
            'pct_demand': round((d2 - d1) / d1 * 100, 2) if d1 > 0 else None,
            'point_elasticity': pe,
        })

    # ── Revenue-maximising price hint ────────────────────────────────
    optimal_price = None
    if revenues and prices:
        max_rev_idx = revenues.index(max(revenues))
        optimal_price = prices[max_rev_idx]

    # ── Summary stats ─────────────────────────────────────────────────
    total_units   = sum(demands)
    total_revenue = sum(revenues)
    avg_price     = sum(p * d for p, d in zip(prices, demands)) / total_units if total_units else 0
    price_range   = f"₾{min(prices):.2f} – ₾{max(prices):.2f}" if prices else "N/A"

    context = {
        # Dropdowns
        'all_groups':    all_groups,
        'all_products':  all_products,
        'all_locations': all_locations,

        # Selected filters
        'selected_group':    selected_group,
        'selected_product':  selected_product,
        'price_type':        price_type,
        'demand_metric':     demand_metric,
        'selected_location': selected_location,

        # Elasticity result
        'elasticity':        elasticity,
        'r_squared':         r_squared,
        'elasticity_label':  elasticity_label,
        'elasticity_color':  elasticity_color,
        'elasticity_advice': elasticity_advice,
        'optimal_price':     optimal_price,

        # Chart data (JSON)
        'chart_prices':   json.dumps(prices),
        'chart_demands':  json.dumps(demands),
        'chart_revenues': json.dumps(revenues),
        'chart_fitted':   json.dumps(fitted),
        'chart_labels':   json.dumps([f"₾{p:.2f}" for p in prices]),

        # Table
        'price_demand_rows': rows,
        'price_field':       price_field,
        'point_elasticities': point_elasticities,

        # Summary
        'total_units':    round(total_units, 0),
        'total_revenue':  round(total_revenue, 2),
        'avg_price':      round(avg_price, 2),
        'price_range':    price_range,
        'row_count':      len(rows),
    }

    return render(request, 'admin_main_only_me.html', context)

@login_required
def client(request):
    # ── Auth / location guard ────────────────────────────────────────────────
    try:
        profile  = request.user.profile
        is_admin = profile.is_admin
    except UserProfile.DoesNotExist:
        is_admin = False

    # ── Filters ──────────────────────────────────────────────────────────────
    selected_location = request.GET.get("location", "all")
    selected_segment  = request.GET.get("segment",  "all")
    selected_year     = request.GET.get("year",     "2025")
    selected_prodg    = request.GET.get("prodg",    "all")

    try:
        year_int = int(selected_year)
    except ValueError:
        year_int = 2025

    # Base queryset
    qs = Sales.objects.filter(cd__year=year_int).exclude(Gvari__startswith='ფიზიკური პირი')
    if not is_admin:
        try:
            allowed = profile.get_allowed_locations()
            if allowed:
                qs = qs.filter(un__in=allowed)
        except Exception:
            pass

    # Dropdown options (always from full table, no year/location filter)
    all_locations = list(
        Sales.objects.values_list("un", flat=True).distinct().order_by("un")
    )
    all_segments = list(
        Sales.objects.values_list("Segment", flat=True)
        .distinct().exclude(Segment__isnull=True).order_by("Segment")
    )
    all_prodgs = list(
        Sales.objects.values_list("prodg", flat=True)
        .distinct().exclude(prodg__isnull=True).order_by("prodg")
    )

    # Apply optional filters
    if selected_location != "all":
        qs = qs.filter(un=selected_location)
    if selected_segment != "all":
        qs = qs.filter(Segment=selected_segment)
    if selected_prodg != "all":
        qs = qs.filter(prodg=selected_prodg)

    # ── 1. SEGMENT OVERVIEW ──────────────────────────────────────────────────
    segment_stats = (
        qs.exclude(Segment__isnull=True)
        .values("Segment")
        .annotate(
            revenue=Sum("tanxa"),
            transactions=Count("zedd", distinct=True),
            customers=Count("IdGvari", distinct=True),
            avg_basket=Avg("tanxa"),
        )
        .order_by("-revenue")
    )
    segment_data = list(segment_stats)
    total_rev = sum(s["revenue"] or 0 for s in segment_data)
    for s in segment_data:
        s["revenue"]    = round(s["revenue"]    or 0, 2)
        s["avg_basket"] = round(s["avg_basket"] or 0, 2)
        s["rev_share"]  = round((s["revenue"] / total_rev * 100) if total_rev else 0, 1)

    seg_data_json = json.dumps([
        {
            "label":        s["Segment"],
            "revenue":      s["revenue"],
            "avg_basket":   s["avg_basket"],
            "transactions": s["transactions"],
            "customers":    s["customers"],
            "rev_share":    s["rev_share"],
        }
        for s in segment_data
    ])

    # ── 2. MONTHLY TREND BY SEGMENT ──────────────────────────────────────────
    monthly_segment = (
        qs.exclude(Segment__isnull=True)
        .annotate(month=TruncMonth("cd"))
        .values("month", "Segment")
        .annotate(revenue=Sum("tanxa"))
        .order_by("month", "Segment")
    )
    seg_months = sorted(set(
        r["month"].strftime("%Y-%m") for r in monthly_segment if r["month"]
    ))
    seg_names = sorted(set(r["Segment"] for r in monthly_segment))
    seg_colors = {
        "Passionate":          "#667eea",
        "Passionate +":        "#a855f7",
        "Regular":             "#10b981",
        "One timer":           "#fb923c",
        "Inactive customers":  "#ef4444",
    }
    seg_monthly_datasets = []
    for seg in seg_names:
        rev_by_month = {
            r["month"].strftime("%Y-%m"): round(r["revenue"] or 0, 2)
            for r in monthly_segment
            if r["Segment"] == seg and r["month"]
        }
        seg_monthly_datasets.append({
            "label":           seg,
            "data":            [rev_by_month.get(m, 0) for m in seg_months],
            "borderColor":     seg_colors.get(seg, "#94a3b8"),
            "backgroundColor": seg_colors.get(seg, "#94a3b8") + "22",
            "tension":         0.4,
            "fill":            False,
            "pointRadius":     3,
        })

    # ── 3. LOCATION × SEGMENT HEATMAP ────────────────────────────────────────
    loc_seg = (
        qs.exclude(Segment__isnull=True)
        .values("un", "Segment")
        .annotate(customers=Count("IdGvari", distinct=True))
        .order_by("un", "Segment")
    )
    locations_for_heatmap = sorted(set(r["un"] for r in loc_seg if r["un"]))
    segments_for_heatmap  = sorted(set(r["Segment"] for r in loc_seg))
    heatmap_matrix = []
    for seg in segments_for_heatmap:
        row_vals = []
        for loc in locations_for_heatmap:
            val = next(
                (r["customers"] for r in loc_seg
                 if r["Segment"] == seg and r["un"] == loc),
                0,
            )
            row_vals.append(val)
        heatmap_matrix.append({"segment": seg, "values": row_vals})

    # ── 4. LOYAL / NAMED CUSTOMER ANALYSIS ───────────────────────────────────
    named_qs = qs.exclude(tanam__isnull=True).exclude(tanam="").exclude()
    top_customers = (
        named_qs.values("IdGvari", "tanam", "Segment", "Gvari")
        .annotate(
            revenue=Sum("tanxa"),
            visits=Count("zedd", distinct=True),
            avg_basket=Avg("tanxa"),
            last_visit=Max("cd"),
            first_visit=Min("cd"),
            categories=Count("prodg", distinct=True),
        )
        .order_by("-revenue")[:50]
    )
    today = timezone.now().date()
    top_customers_list = []
    for c in top_customers:
        last       = c["last_visit"].date()  if c["last_visit"]  else None
        first      = c["first_visit"].date() if c["first_visit"] else None
        days_since = (today - last).days  if last  else None
        tenure     = (today - first).days if first else None
        top_customers_list.append({
            "id":          c["IdGvari"],
            "name":        c["Gvari"]    or "",
            "employee":    c["tanam"]    or "",
            "segment":     c["Segment"] or "",
            "revenue":     round(c["revenue"]    or 0, 2),
            "visits":      c["visits"],
            "avg_basket":  round(c["avg_basket"] or 0, 2),
            "last_visit":  str(last) if last else "",
            "days_since":  days_since,
            "tenure_days": tenure,
            "categories":  c["categories"],
            "recency_flag": (
                "hot"     if days_since is not None and days_since <= 30  else
                "warm"    if days_since is not None and days_since <= 90  else
                "cold"    if days_since is not None and days_since <= 180 else
                "churned"
            ),
        })

    # ── 5. CATEGORY AFFINITY ─────────────────────────────────────────────────
    category_affinity = (
        named_qs.values("Segment", "prodg")
        .annotate(revenue=Sum("tanxa"), txns=Count("zedd", distinct=True))
        .order_by("Segment", "-revenue")
    )
    affinity_by_segment = defaultdict(list)
    for row in category_affinity:
        affinity_by_segment[row["Segment"]].append({
            "category": row["prodg"] or "",
            "revenue":  round(row["revenue"] or 0, 2),
            "txns":     row["txns"],
        })
    affinity_by_segment = dict(affinity_by_segment)

    # ── 6. LOCATION LOYALTY ───────────────────────────────────────────────────
    location_loyalty = (
        qs.exclude(tanam__isnull=True).exclude(tanam="")
        .values("un")
        .annotate(
            loyal_customers=Count("IdGvari", distinct=True),
            loyal_revenue=Sum("tanxa"),
            avg_visits=Avg("raod"),
        )
        .order_by("-loyal_revenue")
    )
    location_loyalty_list = []
    for l in location_loyalty:
        location_loyalty_list.append({
            "un":              l["un"] or "",
            "loyal_customers": l["loyal_customers"],
            "loyal_revenue":   round(l["loyal_revenue"] or 0, 2),
            "avg_visits":      round(l["avg_visits"]    or 0, 2),
        })

    loc_data_json = json.dumps([
        {"label": l["un"], "revenue": l["loyal_revenue"], "customers": l["loyal_customers"]}
        for l in location_loyalty_list
    ])

    # ── 7. RECENCY DISTRIBUTION ───────────────────────────────────────────────
    recency_dist = {"hot": 0, "warm": 0, "cold": 0, "churned": 0}
    for c in top_customers_list:
        recency_dist[c["recency_flag"]] += 1

    # ── 8. PURCHASE FREQUENCY ─────────────────────────────────────────────────
    freq_qs = (
        named_qs.values("IdGvari")
        .annotate(visits=Count("zedd", distinct=True))
    )
    freq_buckets = {"1": 0, "2-3": 0, "4-6": 0, "7-12": 0, "13+": 0}
    for r in freq_qs:
        v = r["visits"]
        if   v == 1:  freq_buckets["1"]    += 1
        elif v <= 3:  freq_buckets["2-3"]  += 1
        elif v <= 6:  freq_buckets["4-6"]  += 1
        elif v <= 12: freq_buckets["7-12"] += 1
        else:         freq_buckets["13+"]  += 1

    # ── 9. WEEKDAY PATTERN ────────────────────────────────────────────────────
    weekday_qs = (
        qs.exclude(Segment__isnull=True)
        .annotate(wd=ExtractWeekDay("cd"))
        .values("wd", "Segment")
        .annotate(revenue=Sum("tanxa"))
        .order_by("wd")
    )
    wd_labels   = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    wd_datasets = []
    for seg in seg_names:
        rev_by_wd = {
            r["wd"]: round(r["revenue"] or 0, 2)
            for r in weekday_qs if r["Segment"] == seg
        }
        wd_datasets.append({
            "label":           seg,
            "data":            [rev_by_wd.get(i, 0) for i in range(1, 8)],
            "backgroundColor": seg_colors.get(seg, "#94a3b8") + "cc",
            "borderRadius":    4,
        })

    # ── CONTEXT ───────────────────────────────────────────────────────────────
    context = {
        "is_admin":             is_admin,
        "user_locations_count": len(all_locations),

        "all_locations":     all_locations,
        "all_segments":      all_segments,
        "all_prodgs":        all_prodgs,
        "selected_location": selected_location,
        "selected_segment":  selected_segment,
        "selected_year":     selected_year,
        "selected_prodg":    selected_prodg,

        # Template rendering
        "segment_data":       segment_data,
        "total_revenue":      round(total_rev, 2),
        "top_customers":      top_customers_list,
        "location_loyalty":   location_loyalty_list,
        "heatmap_segments":   segments_for_heatmap,

        # JSON for JS — ALL data passed this way, no Django loops in <script>
        "seg_data_json":              seg_data_json,
        "loc_data_json":              loc_data_json,
        "seg_months_json":            json.dumps(seg_months),
        "seg_datasets_json":          json.dumps(seg_monthly_datasets),
        "heatmap_locations_json":     json.dumps(locations_for_heatmap),
        "heatmap_matrix_json":        json.dumps(heatmap_matrix),
        "affinity_by_segment_json":   json.dumps(affinity_by_segment),
        "seg_names_json":             json.dumps(seg_names),
        "recency_dist_json":          json.dumps(recency_dist),
        "freq_buckets_json":          json.dumps(freq_buckets),
        "wd_labels_json":             json.dumps(wd_labels),
        "wd_datasets_json":           json.dumps(wd_datasets),
    }

    return render(request, "client_side.html", context)

def compute_metrics(merged: pd.DataFrame) -> dict:
    """
    Returns accuracy metrics where actual data exists.
    MAE, RMSE, MAPE, SMAPE, R², Bias, coverage, error stats.
    """
    has_actual = merged.dropna(subset=["actual"])
    if has_actual.empty:
        return {}

    y        = has_actual["actual"]
    yhat     = has_actual["yhat"]
    error    = y - yhat          # positive = under-forecast
    n        = len(has_actual)

    mae   = error.abs().mean()
    rmse  = (error ** 2).mean() ** 0.5
    mape  = (error.abs() / y.replace(0, pd.NA)).mean() * 100

    # Symmetric MAPE — less sensitive to near-zero actuals
    smape = (2 * error.abs() / (y.abs() + yhat.abs()).replace(0, pd.NA)).mean() * 100

    # R² (coefficient of determination)
    ss_res = (error ** 2).sum()
    ss_tot = ((y - y.mean()) ** 2).sum()
    r2 = 1 - ss_res / ss_tot if ss_tot != 0 else float("nan")

    # Bias (mean signed error) — positive means model under-forecasts
    bias      = error.mean()
    bias_pct  = (bias / y.mean()) * 100 if y.mean() != 0 else float("nan")

    # Directional accuracy — did forecast move in same direction as actual day-over-day?
    if n > 1:
        dir_acc = (
            ((y.diff() > 0) == (yhat.diff() > 0)).iloc[1:].mean() * 100
        )
    else:
        dir_acc = float("nan")

    # Cumulative sums (for tracking total gap)
    cum_actual   = round(float(y.sum()), 2)
    cum_forecast = round(float(yhat.sum()), 2)
    cum_error    = round(cum_actual - cum_forecast, 2)
    cum_error_pct = round((cum_error / cum_actual) * 100, 2) if cum_actual != 0 else float("nan")

    # CI coverage
    in_ci = (
        (y >= has_actual["yhat_lower"]) & (y <= has_actual["yhat_upper"])
    ).mean() * 100

    # Max single-day absolute error
    max_ae     = round(float(error.abs().max()), 2)
    max_ae_day = has_actual.loc[error.abs().idxmax(), "ds"].strftime("%Y-%m-%d") if not error.empty else None

    # Percentile errors
    p50_err = round(float(error.abs().quantile(0.50)), 2)
    p90_err = round(float(error.abs().quantile(0.90)), 2)

    return {
        # Core
        "mae":           round(float(mae), 2),
        "rmse":          round(float(rmse), 2),
        "mape":          round(float(mape), 2),
        "smape":         round(float(smape), 2),
        "r2":            round(float(r2), 4),
        # Bias
        "bias":          round(float(bias), 2),
        "bias_pct":      round(float(bias_pct), 2),
        # Cumulative
        "cum_actual":    cum_actual,
        "cum_forecast":  cum_forecast,
        "cum_error":     cum_error,
        "cum_error_pct": cum_error_pct,
        # Distribution
        "max_ae":        max_ae,
        "max_ae_day":    max_ae_day,
        "p50_err":       p50_err,
        "p90_err":       p90_err,
        # Direction & CI
        "dir_acc":       round(float(dir_acc), 1) if not pd.isna(dir_acc) else None,
        "coverage":      round(float(in_ci), 2),
        "n_actual":      int(n),
        # Legacy key kept for KPI card
        "total_actual":  cum_actual,
    }

@login_required
def Forecast_fb(request):
    # 1. Load forecast file
    path = os.path.join(
        settings.BASE_DIR, "sales_app", "data", "Forecast_June_Base.xlsx"
    )
    forecast_df = pd.read_excel(path, engine="openpyxl", sheet_name="Sheet1")

    # Normalise column names
    forecast_df.columns = forecast_df.columns.str.strip()
    forecast_df["ds"] = pd.to_datetime(forecast_df["ds"]).dt.normalize()

    # 2. Respect user's allowed locations
    try:
        profile  = request.user.profile
        allowed  = profile.get_allowed_locations()
        is_admin = profile.is_admin
    except Exception:
        allowed  = []
        is_admin = True

    all_locations = forecast_df["location"].unique().tolist()

    if not is_admin and allowed:
        forecast_df = forecast_df[forecast_df["location"].isin(allowed)]

    locations = forecast_df["location"].unique().tolist()

    # 3. Pull actual sales from DB for the same date range
    date_min = forecast_df["ds"].min()
    date_max = forecast_df["ds"].max()

    sales_qs = Sales.objects.filter(
        cd__date__gte=date_min,
        cd__date__lte=date_max,
    )
    if not is_admin and allowed:
        sales_qs = sales_qs.filter(un__in=allowed)

    sales_df = pd.DataFrame(list(sales_qs.values("un", "cd", "tanxa")))

    # 4. Aggregate actuals: daily revenue per location
    if not sales_df.empty:
        sales_df["cd"] = pd.to_datetime(sales_df["cd"]).dt.normalize()
        actuals = (
            sales_df.groupby(["un", "cd"])["tanxa"]
            .sum()
            .reset_index()
            .rename(columns={"un": "location", "cd": "ds", "tanxa": "actual"})
        )
    else:
        actuals = pd.DataFrame(columns=["location", "ds", "actual"])

    # 5. Merge forecast ← actuals
    actuals["location"] = actuals["location"].str.strip()
    merged = forecast_df.merge(actuals, on=["location", "ds"], how="left")
    merged["ds_str"] = merged["ds"].dt.strftime("%Y-%m-%d")

    # 6. Build per-location chart data + metrics
    chart_data   = {}
    metrics_data = {}
    today = date.today()


    for loc in locations:
        loc_df = merged[merged["location"] == loc].sort_values("ds")

        # Daily forecast error series (for error chart)
        daily_error     = (loc_df["actual"] - loc_df["yhat"]).where(loc_df["actual"].notna(), None)
        daily_error_pct = (daily_error / loc_df["yhat"].replace(0, pd.NA) * 100).where(loc_df["actual"].notna(), None)

        chart_data[loc] = {
            "dates":           loc_df["ds_str"].tolist(),
            "yhat":            loc_df["yhat"].round(2).tolist(),
            "yhat_lower":      loc_df["yhat_lower"].round(2).tolist(),
            "yhat_upper":      loc_df["yhat_upper"].round(2).tolist(),
            "actual":          loc_df["actual"].where(loc_df["actual"].notna(), None).tolist(),
            "daily_error": [round(v, 2) if v is not None else None for v in daily_error],
            "daily_error_pct": daily_error_pct.round(1).tolist(),
        }
        metrics_data[loc] = compute_metrics(loc_df)

    # 7. Summary stats across all locations
    total_forecast = merged.groupby("ds_str")["yhat"].sum().reset_index()
    total_actual   = merged.groupby("ds_str")["actual"].sum().reset_index()

    # Running cumulative series for the summary chart
    cum_f = total_forecast["yhat"].cumsum().round(2).tolist()
    cum_a_series = total_actual["actual"].where(total_actual["actual"].notna(), None)
    cum_a = cum_a_series.cumsum().where(cum_a_series.notna(), None).round(2).tolist()

    summary_chart = {
        "dates":       total_forecast["ds_str"].tolist(),
        "yhat":        total_forecast["yhat"].round(2).tolist(),
        "actual":      total_actual["actual"].where(total_actual["actual"].notna(), None).tolist(),
        "cum_forecast": cum_f,
        "cum_actual":   cum_a,
    }

    # Scalar KPIs
    total_forecast_sum = merged["yhat"].sum()
    total_actual_sum   = merged["actual"].sum(skipna=True)
    avg_mape = (
        pd.Series([m.get("mape") for m in metrics_data.values() if m.get("mape") is not None])
        .mean()
    )
    avg_r2 = (
        pd.Series([m.get("r2") for m in metrics_data.values() if m.get("r2") is not None])
        .mean()
    )

    context = {
        "locations":          locations,
        "chart_data_json":    mark_safe(json.dumps(chart_data,    ensure_ascii=False)),
        "metrics_data_json":  mark_safe(json.dumps(metrics_data,  ensure_ascii=False)),
        "summary_chart_json": mark_safe(json.dumps(summary_chart, ensure_ascii=False)),
        "total_forecast":     round(total_forecast_sum, 0),
        "total_actual":       round(total_actual_sum, 0) if total_actual_sum else None,
        "avg_mape":           round(avg_mape, 1) if pd.notna(avg_mape) else None,
        "avg_r2":             round(avg_r2, 3)  if pd.notna(avg_r2)   else None,
        "n_locations":        len(locations),
        "is_admin":           is_admin,
    }

    return render(request, "ForecastMain.html", context)