import json
import io
from datetime import datetime
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required  
from django.db import connection
from .llm import generate_sql
from .executor import run_query
from functools import wraps  #

# ↓ paste the decorator here
def admin_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        try:
            if not request.user.profile.is_admin:
                return JsonResponse({"error": "Access denied"}, status=403)
        except Exception:
            return JsonResponse({"error": "Access denied"}, status=403)
        return view_func(request, *args, **kwargs)
    return wrapper

@admin_required
@require_POST
def ai_query(request):
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    user_prompt = body.get("prompt", "").strip()
    if not user_prompt:
        return JsonResponse({"error": "No prompt provided"}, status=400)

    try:
        profile = request.user.profile
        allowed_locations = profile.get_allowed_locations()
    except Exception:
        return JsonResponse({"error": "User profile not found"}, status=403)

    sql = generate_sql(user_prompt, allowed_locations)
    result = run_query(sql)
    return JsonResponse(result)


@admin_required
@require_POST
def export_excel(request):
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    sql = body.get("sql", "").strip()
    if not sql:
        return JsonResponse({"error": "No SQL provided"}, status=400)

    # Safety: only allow SELECT
    if not re.match(r'^\s*(SELECT|WITH)\b', sql.upper()):
        return JsonResponse({"error": "Only SELECT queries are allowed"}, status=400)

    try:
        with connection.cursor() as cursor:
            cursor.execute(sql)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
    except Exception as e:
        return JsonResponse({"error": f"Query failed: {str(e)}"}, status=500)

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    # Stream response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_{timestamp}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response