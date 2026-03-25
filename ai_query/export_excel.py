import io
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.db import connection


def export_query_to_excel(sql: str, filename_prefix: str = "export") -> HttpResponse:
    """
    Run a SQL SELECT query and return results as an Excel file download.
    """
    with connection.cursor() as cursor:
        cursor.execute(sql)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # --- Header row styling ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Data rows ---
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Auto-fit column widths ---
    for col_cells in ws.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_length + 4, 50)

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Write to response ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response